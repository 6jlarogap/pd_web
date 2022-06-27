# yalta-export.py
#
# Экспорт данных для Ялты
#
# Запуск из ./manage.py shell :
# exec(open('../contrib/yalta-export.py').read())

UGH_PK = 394
FNAME = 'report.xlsx'

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from burials.models import Burial

def main():
    fname = FNAME
    try:
        os.unlink(fname)
    except OSError:
        pass

    align_right = Alignment(horizontal="right", vertical="center", indent=1)
    align_left = Alignment(horizontal="left", vertical="center", indent=1)
    align_сenter = Alignment(horizontal="center", vertical="center")
    font_header  = Font(name='Arial', b=True, size=12)

    book = Workbook()
    sheet = book.active
    sheet.title = 'Экспорт'
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 50
    sheet.column_dimensions['D'].width = 50
    for c in range(1, 5):
        sheet.cell(1, c).alignment = align_сenter
    row = 1
    title = sheet.cell(row, 1)
    title.value = '№ п/п'
    title.font = font_header
    title = sheet.cell(row, 2)
    title.value = 'Рег.'
    title.font = font_header
    title = sheet.cell(row, 3)
    title.value = 'ФИО умершего'
    title.font = font_header
    title = sheet.cell(row, 4)
    title.value = 'ФИО ответственного'
    title.font = font_header
    
    for b in Burial.objects.filter(
            ugh__pk=UGH_PK,
            status=Burial.STATUS_CLOSED,
            place__isnull=False,
        ).select_related(
            'deadman',
            'place__responsible',
        ).order_by(
            'pk'):
        row += 1
        cell_pk = sheet.cell(row, 1)
        cell_pk.value = str(b.pk)
        cell_pk.alignment = align_right

        cell_an = sheet.cell(row, 2)
        cell_an.value = str(b.account_number) or ''
        cell_an.alignment = align_left

        cell_deadman = sheet.cell(row, 3)
        cell_deadman.value = b.deadman and b.deadman.last_name and b.deadman.full_human_name() or ''
        cell_deadman.alignment = align_left

        cell_responsible = sheet.cell(row, 4)
        cell_responsible.value = b.place.responsible and b.place.responsible.last_name and b.place.responsible.full_human_name() or ''
        cell_responsible.alignment = align_left

    book.save(fname)

main()

