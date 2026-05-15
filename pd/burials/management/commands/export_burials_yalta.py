# export_burials_yalta.py
#
# Запуск: ./manage.py export_burials_yalta <ugh_pk> <output.xlsx> <date_from> <date_to>
# Даты в формате ДД.ММ.ГГГГ
#
# Экспорт захоронений Ялты в Excel (колонки = таблица поиска захоронений)

import datetime
import os

from django.core.management.base import BaseCommand, CommandError

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from burials.models import Burial
from users.models import Org


DATE_FORMAT = '%d.%m.%Y'

HEADERS = [
    '№ п/п',
    'Номер в книге',
    'Заявитель',
    'Кладбище',
    'Участок',
    'Ряд',
    'Место',
    'Усопший',
    'Дата рождения',
    'Дата смерти',
    'Дата захоронения',
    'Статус',
    'Ответственный',
]

COL_WIDTHS = [6, 16, 30, 20, 12, 8, 10, 30, 14, 14, 18, 14, 30]


class Command(BaseCommand):

    help = 'Export burials for Yalta UGH to Excel (burial_list columns)'

    def add_arguments(self, parser):
        parser.add_argument('ugh_pk', type=int)
        parser.add_argument('output', type=str)
        parser.add_argument('date_from', type=str, help='DD.MM.YYYY')
        parser.add_argument('date_to', type=str, help='DD.MM.YYYY')

    def handle(self, *args, **kwargs):
        ugh_pk = kwargs['ugh_pk']
        output = kwargs['output']
        date_from_str = kwargs['date_from']
        date_to_str = kwargs['date_to']

        try:
            date_from = datetime.datetime.strptime(date_from_str, DATE_FORMAT).date()
            date_to = datetime.datetime.strptime(date_to_str, DATE_FORMAT).date()
        except ValueError as e:
            raise CommandError('Bad date format: %s' % e)

        try:
            ugh = Org.objects.get(pk=ugh_pk)
        except Org.DoesNotExist:
            raise CommandError('Org pk=%s not found' % ugh_pk)

        qs = Burial.objects.filter(
            ugh=ugh,
            status=Burial.STATUS_CLOSED,
            annulated=False,
            fact_date__gte=date_from,
            fact_date__lte=date_to,
        ).select_related(
            'cemetery',
            'area',
            'deadman',
            'applicant',
            'applicant_organization',
            'responsible',
            'place__responsible',
        ).order_by('fact_date', 'pk')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Захоронения'

        font_header = Font(name='Arial', bold=True, size=10)
        font_body = Font(name='Arial', size=10)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

        for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = font_header
            cell.alignment = align_center
            ws.column_dimensions[cell.column_letter].width = width
        ws.row_dimensions[1].height = 30

        n = 0
        for b in qs.iterator(chunk_size=200):
            n += 1
            row = n + 1

            # Заявитель
            if b.applicant_organization:
                applicant = str(b.applicant_organization)
            elif b.applicant:
                applicant = b.applicant.full_human_name()
            else:
                applicant = ''

            # Усопший
            if b.deadman:
                deadman_fio = b.deadman.full_human_name()
                birth_date = b.deadman.birth_date and b.deadman.birth_date.strftime(DATE_FORMAT) or ''
                death_date = b.deadman.death_date and b.deadman.death_date.strftime(DATE_FORMAT) or ''
            else:
                deadman_fio = ''
                birth_date = ''
                death_date = ''

            # Дата захоронения
            fact_date = b.fact_date and b.fact_date.strftime(DATE_FORMAT) or ''

            # Ответственный
            responsible = b.get_responsible()
            responsible_fio = responsible.full_human_name() if responsible else ''

            values = [
                n,
                b.account_number or '',
                applicant,
                b.cemetery.name if b.cemetery else '',
                b.area.name if b.area else '',
                b.row or '',
                b.place_number or '',
                deadman_fio,
                birth_date,
                death_date,
                fact_date,
                b.get_status_display(),
                responsible_fio,
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.font = font_body
                cell.alignment = align_left if col_idx > 1 else align_center

        ws.freeze_panes = 'A2'

        try:
            wb.save(output)
        except OSError as e:
            raise CommandError('Cannot save %s: %s' % (output, e))

        self.stdout.write(self.style.SUCCESS(
            'Exported %d burials to %s' % (n, output)
        ))
