import datetime
import decimal
import json
import random, math
import string
import re, os, tempfile
import hashlib
import pytils

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from django.conf import settings
from django.contrib import messages
from django.urls import reverse
from django.core.exceptions import ValidationError, PermissionDenied
from django.db import transaction
from django.db.models import F
from django.db.models.aggregates import Count, Sum
from django.db.models.query_utils import Q
from django.http import HttpResponse, Http404
from django.shortcuts import redirect, render
from django.template.loader import render_to_string
from django.views.generic.base import View
from django.views.generic.detail import DetailView
from django.views.generic.edit import CreateView, UpdateView, FormView
from django.views.generic.list import ListView
from django.utils.translation import ugettext_lazy as _
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.shortcuts import get_object_or_404

from LatLon23 import LatLon, Latitude, Longitude

from logs.models import write_log
from geo.models import Location
from burials.forms import AddOrgForm, AddAgentForm, AddDoverForm, AddDocTypeForm
from burials.models import Burial, Place, Area, Cemetery, OrderPlace
from burials.views import TradeCemeteriesMixin
from users.models import CustomerProfile, Org, ProfileLORU, Store, OrgWebPay, \
                         is_trade_user, is_supervisor, is_cabinet_user, \
                         PermitIfTrade, PermitIfCabinet, PermitIfTradeOrCabinet, \
                         get_profile
from billing.models import Rate
from orders.forms import ProductForm, OrderForm, OrderItemFormset, CoffinForm, CatafalqueForm, \
                         AddInfoForm, OrderSearchForm, OrderBurialForm, \
                         OrderServicesForm, ProductXlsxreportForm
from orders.models import Product, Order, OrderItem, ProductCategory, \
                          Service, Measure, OrgService, OrgServicePrice, ServiceItem, OrderComment, \
                          Route, ResultFile, OrderWebPay
from persons.models import CustomPlace, AlivePerson, CustomPerson, OrderDeadPerson
from pd.forms import CommentForm
from pd.views import PaginateListView, RequestToFormMixin, ServiceException, get_front_end_url, get_host_url
from reports.models import make_report

from rest_framework import viewsets
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.renderers import StaticHTMLRenderer

from orders.serializers import ProductCategorySerializer, ProductsSerializer, ProductsOptSerializer, \
                               ProductInfoSerializer, OptOrdersSerializer, OptOrderInfoSerializer, \
                               ProductEditSerializer, ServiceSerializer, OrgServiceSerializer, \
                               ServiceOrderSerializer, OrderCommentsSerializer, ServiceOrderDetailSerializer, \
                               OrderResultsSerializer, ProductCategory2Serializer, LoruOrderSerializer

from rest_api.fields import UnclearDateFieldSerializer, UnclearDateFieldMixin
from pd.utils import EmailMessage, str_to_bool_or_None, get_image, is_video, re_search
from pd.models import validate_phone_as_number, CheckLifeDatesMixin, SafeDeleteMixin

from sms_service.utils import send_sms
from functools import reduce

class ProductCategoryQsMixin(object):
    
    def product_category_qs(self):
        category_ids_got = self.request.GET.getlist('filter[category]')
        category_ids = []
        for category_id in category_ids_got:
            try:
                # может быть '', 'null'
                category_ids.append(int(category_id))
            except ValueError:
                pass
        if category_ids:
            return Q(productcategory__pk__in=category_ids)
        else:
            return Q()

class LORURequiredMixin:
    def is_loru(self, request):
        if not request.user.is_authenticated:
            return False
        if not getattr(self.request.user, 'profile', None):
            return False
        if not self.request.user.profile.is_loru():
            return False
        return True

    def dispatch(self, request, *args, **kwargs):
        self.request = request
        return self.is_loru(request) and View.dispatch(self, request, *args, **kwargs) or redirect('/')

class ProductList(LORURequiredMixin, ListView):
    template_name = 'product_list.html'
    model = Product

    def get_queryset(self):
        return Product.objects.filter(loru=self.request.user.profile.org)
    
manage_products = ProductList.as_view()

class ProductCreate(LORURequiredMixin, RequestToFormMixin, CreateView):
    template_name = 'product_edit.html'
    form_class = ProductForm

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.loru = self.request.user.profile.org
        self.object.save()
        if not self.object.sku or not self.object.sku.strip():
            self.object.sku = str(self.object.pk)
            self.object.save()
        write_log(self.request, self.object, _('Создание: %s') % self.object.name)
        msg = _("<a href='%(manage_products_edit)s'>Товар %(name)s</a> создан") % dict(
            manage_products_edit=reverse('manage_products_edit', args=[self.object.pk]),
            name=self.object.name,
        )
        messages.success(self.request, msg)
        return redirect('manage_products')

manage_products_create = ProductCreate.as_view()

class ProductEdit(LORURequiredMixin, RequestToFormMixin, UpdateView):
    template_name = 'product_edit.html'
    form_class = ProductForm

    def get_queryset(self):
        return Product.objects.filter(loru=self.request.user.profile.org)

    def form_valid(self, form):
        self.object = form.save(commit=False)
        if not self.object.sku or not self.object.sku.strip():
            self.object.sku = str(self.object.pk)
        self.object.save()
        write_log(self.request, self.object, _('Изменение: %s') % self.object.name)
        msg = _("<a href='%(manage_products_edit)s'>Товар %(name)s</a> изменен") % dict(
            manage_products_edit=reverse('manage_products_edit', args=[self.object.pk]),
            name=self.object.name,
        )
        messages.success(self.request, msg)
        return redirect('manage_products')

manage_products_edit = ProductEdit.as_view()

class OrderList(LORURequiredMixin, PaginateListView):
    template_name = 'order_list.html'
    model = Order

    def __init__(self, *args, **kwargs):
        super(OrderList, self).__init__(*args, **kwargs)
        self.SORT_DEFAULT = '-order_date'
        
    def get_queryset(self):
        if not self.request.GET:
            return Order.objects.none().order_by('-pk')

        return self.filtered_orders()

    def filtered_orders(self):
        q = Q(loru=self.request.user.profile.org) | \
            Q(type=Order.TYPE_TRADE, applicant_organization=self.request.user.profile.org)
        orders = Order.objects.filter(q).distinct()
                # .annotate(item_count=Count('orderitem'))  # мы не показываем в таблице кол-во товаров,
                                                            # к тому же это резко замедляет поиск

        form = self.get_form()
        if form.data and form.is_valid():
            fio_string = form.cleaned_data['fio_order_deadman']
            if fio_string:
                search_by =  [
                    'burial__deadman__last_name__iregex',
                    'burial__deadman__first_name__iregex',
                    'burial__deadman__middle_name__iregex'
                ]
                q_burial_deadman = self.q_by_name(search_by=search_by, name_string=fio_string)
                search_by =  [
                    'orderdeadperson__last_name__iregex',
                    'orderdeadperson__first_name__iregex',
                    'orderdeadperson__middle_name__iregex'
                ]
                q_orderdeadperson = self.q_by_name(search_by=search_by, name_string=fio_string)
                orders = orders.filter(q_burial_deadman | q_orderdeadperson)
            birth_date_from = form.cleaned_data['birth_date_from']
            if birth_date_from:
                orders = orders.filter(
                    Q(burial__deadman__birth_date__gte=birth_date_from) | \
                    Q(orderdeadperson__birth_date__gte=birth_date_from)
                )
            birth_date_to = form.cleaned_data['birth_date_to']
            if birth_date_to:
                birth_date_to = birth_date_to + datetime.timedelta(days=1)
                orders = orders.filter(
                    Q(burial__deadman__birth_date__lt=birth_date_to) | \
                    Q(orderdeadperson__birth_date__lt=birth_date_to)
                )
            death_date_from = form.cleaned_data['death_date_from']
            if death_date_from:
                orders = orders.filter(
                    Q(burial__deadman__death_date__gte=death_date_from) | \
                    Q(orderdeadperson__death_date__gte=death_date_from)
                )
            death_date_to = form.cleaned_data['death_date_to']
            if death_date_to:
                death_date_to = death_date_to + datetime.timedelta(days=1)
                orders = orders.filter(
                    Q(burial__deadman__death_date__lt=death_date_to) | \
                    Q(orderdeadperson__death_date__lt=death_date_to)
                )
            burial_date_from = form.cleaned_data['burial_date_from']
            if burial_date_from:
                orders = orders.filter(
                    Q(burial__plan_date__gte=burial_date_from) | \
                    Q(dt_due__gte=burial_date_from)
                )
            burial_date_to = form.cleaned_data['burial_date_to']
            if burial_date_to:
                burial_date_to = burial_date_to + datetime.timedelta(days=1)
                orders = orders.filter(
                    Q(burial__plan_date__lte=burial_date_to) | \
                    Q(dt_due__lt=burial_date_to)
                )
            if form.cleaned_data['account_number_from']:
                orders = orders.filter(loru_number__gte=form.cleaned_data['account_number_from'])
            if form.cleaned_data['account_number_to']:
                orders = orders.filter(loru_number__lte=form.cleaned_data['account_number_to'])
            if form.cleaned_data['responsible']:
                fio = [re_search(f) for f in form.cleaned_data['responsible'].split()]
                q1r = Q(burial__responsible__isnull=False)
                q2r = Q(burial__place__isnull=False)
                if len(fio) > 2:
                    q1r &= Q(burial__responsible__middle_name__iregex=fio[2])
                    q2r &= Q(burial__place__responsible__middle_name__iregex=fio[2])
                if len(fio) > 1:
                    q1r &= Q(burial__responsible__first_name__iregex=fio[1])
                    q2r &= Q(burial__place__responsible__first_name__iregex=fio[1])
                if len(fio) > 0:
                    q1r &= Q(burial__responsible__last_name__iregex=fio[0])
                    q2r &= Q(burial__place__responsible__last_name__iregex=fio[0])
                    qr = Q(q1r | q2r)
                    orders = orders.filter(qr)
            if form.cleaned_data['cemetery']:
                orders = orders.filter(burial__cemetery__name=form.cleaned_data['cemetery'])
            if form.cleaned_data['area']:
                orders = orders.filter(burial__area__name=form.cleaned_data['area'])
            if form.cleaned_data['row']:
                orders = orders.filter(burial__row=form.cleaned_data['row'])
            if form.cleaned_data['place']:
                orders = orders.filter(burial__place_number=form.cleaned_data['place'])
            if form.cleaned_data['no_last_name']:
                orders = orders.filter(Q(burial__deadman__last_name='') | Q(burial__deadman__last_name__isnull=True))
            if form.cleaned_data['no_responsible']:
                orders = orders.filter(burial__place__responsible__isnull=True)
            if form.cleaned_data['status']:
                orders = orders.filter(burial__status=form.cleaned_data['status'])
            if form.cleaned_data['order_num_from']:
                orders = orders.filter(loru_number__gte=form.cleaned_data['order_num_from'])
            if form.cleaned_data['order_num_to']:
                orders = orders.filter(loru_number__lte=form.cleaned_data['order_num_to'])
            if form.cleaned_data['order_cost_from']:
                orders = orders.filter(cost__gte=form.cleaned_data['order_cost_from'])
            if form.cleaned_data['order_cost_to']:
                orders = orders.filter(cost__lte=form.cleaned_data['order_cost_to'])
            if form.cleaned_data['annulated']:
                orders = orders.filter(annulated=True)
            else:
                orders = orders.filter(annulated=False)
            if form.cleaned_data['burial_num_from']:
                orders = orders.filter(burial__id__gte = form.cleaned_data['burial_num_from'])
            if form.cleaned_data['burial_num_to']:
                orders = orders.filter(burial__id__lte = form.cleaned_data['burial_num_to'])
            if form.cleaned_data['applicant_org']:
                orders = orders.filter(applicant_organization__name__istartswith=form.cleaned_data['applicant_org'])
            if form.cleaned_data['applicant_person']:
                search_by =  [
                    'applicant__last_name__iregex',
                    'applicant__first_name__iregex',
                    'applicant__middle_name__iregex'
                ]
                orders = self.filter_by_name(queryset=orders, search_by=search_by, name_string=form.cleaned_data['applicant_person'])
            if form.cleaned_data['reg_number_from']:
                orders = orders.filter(burial__account_number__gte = form.cleaned_data['reg_number_from'])
            if form.cleaned_data['reg_number_to']:
                orders = orders.filter(burial__account_number__lte= form.cleaned_data['reg_number_to'])
            if form.cleaned_data['burial_container']:
                orders = orders.filter(burial__burial_container=form.cleaned_data['burial_container'])
        else:
            orders = orders.exclude(annulated=True)

        orders_count = orders.count()
        orders = orders.select_related(
            'burial', 'burial__ugh', 'burial__cemetery', 'burial__area', 'burial__responsible',
            'burial__changed_by', 'burial__deadman', 'applicant_organization', 'applicant',
        )

        sort = self.request.GET.get('sort', self.SORT_DEFAULT)
        SORT_FIELDS = {
            'order_date': 'dt',
            '-order_date': '-dt',
            'account_number': 'burial__account_number',
            '-account_number': '-burial__account_number',
            'deadman': 'burial__deadman__last_name',
            '-deadman': '-burial__deadman__last_name',
            'plan_date': 'burial__plan_date',
            '-plan_date': '-burial__plan_date',
            'place': 'burial__place',
            '-place': '-burial__place',
            'cemetery': 'burial__cemetery',
            '-cemetery': '-burial__cemetery',
            'burial': 'burial__pk',
            '-burial': '-burial__pk',
            'applicant': ['applicant', 'applicant_organization'],
            '-applicant': ['-applicant', '-applicant_organization'],
            'order_num': 'loru_number',
            '-order_num': '-loru_number',
            'cost': 'cost',
            '-cost': '-cost',
        }
        s = SORT_FIELDS[sort]
        if not isinstance(s, list):
            s = [s]
        orders = orders.order_by(*s)

        orders.count = lambda: orders_count
        return orders

    def get_form(self):
        return OrderSearchForm(data=self.request.GET or None)

    def q_by_name(self, search_by, name_string):
        import operator
        values = [re_search(f) for f in name_string.split()]
        predicates = list(zip(search_by, values))
        query = [Q(p) for p in predicates]
        q = reduce(operator.and_, query)
        return q

    def filter_by_name(self, queryset, search_by, name_string):
        q = self.q_by_name(search_by, name_string)
        return queryset.filter(q)

    def get(self, request, *args, **kwargs):
        if self.request.GET.get('burial_plan'):
            form = self.get_form()
            no_dates = True
            if form.data and form.is_valid():
                burial_date_from = form.cleaned_data['burial_date_from']
                burial_date_to = form.cleaned_data['burial_date_to']
                no_dates = not (burial_date_from and burial_date_to)
            if no_dates:
                return render(
                    request,
                    'simple_message.html',
                    dict(message=_("Не задан интервал дат захоронения"))
                )
            if burial_date_from > burial_date_to:
                return render(
                    request,
                    'simple_message.html',
                    dict(message=_('"Дата захоронений: с" больше "Дата захоронений: по"'))
                )
            orders = self.filtered_orders()
            if not orders:
                return render(
                    request,
                    'simple_message.html',
                    dict(message=_('Не найдены заказы по выбранным критериям'))
                )
            # Это почистит предыдущий order_by
            orders = orders.order_by('dt_due', 'burial__plan_date')
            cur_date = None
            dates_pre = dict()
            for order in orders:
                order_date = order.dt_due or order.burial.plan_date
                if order_date in dates_pre:
                    dates_pre[order_date].append(order)
                else:
                    dates_pre[order_date] = [order]
            dates = [dict(date=date, orders=dates_pre[date]) for date in sorted(dates_pre.keys())]
            context = dict(
                dates=dates,
                start_date=burial_date_from,
                end_date=burial_date_to,
                print_now=True,
            )
            return render(request, 'reports/burial_plan.html', context)
        if not request.GET and request.user.profile.org.inn == '9103078189':
            form = self.get_form()
            get_attrs = '?'
            for k in list(form.fields.keys()):
                get_attrs += '%s=%s&' % (
                    k,
                    '25' if k == 'per_page' else '',
                )
            get_attrs += "sort=%s" % "-order_num"
            return redirect(reverse('order_list') + get_attrs)
        return super(OrderList, self).get(request, *args, **kwargs)

order_list = OrderList.as_view()

class OrderCreate(LORURequiredMixin, RequestToFormMixin, CreateView):
    template_name = 'order_create.html'
    form_class = OrderForm

    def dispatch(self, request, *args, **kwargs):
        self.request = request
        self.burial = None
        if self.is_loru(request):
            burial_pk = self.request.GET.get('burial') or self.request.POST.get('burial')
            if burial_pk:
                try:
                    self.burial = Burial.objects.get(pk=burial_pk)
                    if not self.burial.can_bind_to_order(self.request.user.profile.org):
                        raise Http404
                except Burial.DoesNotExist:
                    raise Http404
            return View.dispatch(self, request, *args, **kwargs)
        else:
            raise Http404
        
    def get_context_data(self, **kwargs):
        data = super(OrderCreate, self).get_context_data(**kwargs)
        data.update({
            'agent_form': AddAgentForm(prefix='agent'),
            'agent_dover_form': AddDoverForm(prefix='agent_dover'),
            'dover_form': AddDoverForm(prefix='dover'),
            'org_form': AddOrgForm(request=self.request, prefix='org'),
            'doc_type_form': AddDocTypeForm(prefix='doctype'),
        })
        return data

    @transaction.atomic
    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.loru = self.request.user.profile.org
        if self.burial:
            self.object.burial = self.burial
        self.object.save()

        for p in Product.objects.filter(loru=self.request.user.profile.org, default=True, is_archived=False):
            OrderItem.objects.create(order=self.object, product=p)

        write_log(self.request, self.object, _('Заказ сохранен'))
        msg = _("<a href='%(order_edit)s'>Заказ %(pk)s</a> сохранен") % dict(
            order_edit=reverse('order_edit', args=[self.object.pk]),
            pk=self.object.pk,
        )
        messages.success(self.request, msg)
        return redirect('order_products', self.object.pk)

order_create = OrderCreate.as_view()

class OrderEdit(LORURequiredMixin, RequestToFormMixin, UpdateView):
    template_name = 'order_edit_applicant.html'
    form_class = OrderForm

    def get_context_data(self, **kwargs):
        data = super(OrderEdit, self).get_context_data(**kwargs)
        data.update({
            'order': self.get_object(),
            'agent_form': AddAgentForm(prefix='agent'),
            'agent_dover_form': AddDoverForm(prefix='agent_dover'),
            'dover_form': AddDoverForm(prefix='dover'),
            'org_form': AddOrgForm(request=self.request, prefix='org'),
        })
        return data

    def get(self, request, pk, *args, **kwargs):
        self.request = request
        try:
            o = Order.objects.get(loru=self.request.user.profile.org, pk=pk)
        except Order.DoesNotExist:
            raise Http404
        front_end_url = get_front_end_url(request)
        if o.is_type_funeral():
            return redirect("%sloru/own-orders/%s" % (front_end_url, pk))
        elif o.is_type_trade():
            return redirect("%sorder/%s" % (front_end_url, pk))
        elif o.is_type_customer():
            return redirect("%sorders/%s" % (front_end_url, pk))
        # else o.is_type_burial()
        # pass
        if self.request.session.get('order_burial_saved'):
            del self.request.session['order_burial_saved']
            if self.get_object().has_services:
                return redirect('order_services', self.get_object().pk)
        return super(OrderEdit, self).get(request, *args, **kwargs)

    def get_queryset(self):
        return Order.objects.filter(loru=self.request.user.profile.org, type=Order.TYPE_BURIAL)

    def form_valid(self, form):
        self.object = form.save()
        go_next = '_save_next' in self.request.POST

        write_log(self.request, self.object, _('Заказ сохранен'))
        if go_next:
            return redirect('order_products', self.object.pk)
        else:
            msg = _("<a href='%(order_edit)s'>Заказ %(pk)s</a> сохранен") % dict(
                order_edit=reverse('order_edit', args=[self.object.pk]),
                pk=self.object.pk,
            )
            messages.success(self.request, msg)
            return redirect('.')

order_edit = OrderEdit.as_view()

class AjaxProductPrice(LORURequiredMixin, View):
    def get(self, request, *args, **kwargs):
        id = request.GET.get('id')
        if id:
            product = get_object_or_404(Product, pk=id)
            price = product.price
        else:
            price = 0
        return HttpResponse(json.dumps({'price': float(price)}), content_type='application/json')

ajax_product_price = AjaxProductPrice.as_view()

class OrderEditProducts(LORURequiredMixin, View):
    template_name = 'order_edit_products.html'

    def get_queryset(self):
        return Order.objects.filter(loru=self.request.user.profile.org)

    def get_formset(self):
        return OrderItemFormset(request=self.request, data=self.request.POST or None, instance=self.get_object())

    def get_form(self):
        instance = self.get_object()
        price_photo = instance.price_photo()
        form = OrderServicesForm(data=self.request.POST or None)
        if price_photo is None:
            del form.fields['do_photo']
            del form.fields['price_photo']
        else:
            form.initial.update(dict(
                do_photo=instance.has_photo(),
                price_photo=str(price_photo),
            ))
            form.fields['do_photo'].label += ", %s%s. " % (
                price_photo,
                instance.loru.currency.one_char_name(),
            )
        return form

    def get_object(self):
        return self.get_queryset().get(pk=self.kwargs['pk'], type=Order.TYPE_BURIAL)

    def get_context_data(self, **kwargs):
        return {
            'order': self.get_object(),
            'form': self.get_form(),
            'formset': self.get_formset(),
        }

    @transaction.atomic
    def post(self, request, *args, **kwargs):
            
        formset = self.get_formset()
        if formset.is_valid():
            self.object = self.get_object()
            go_next = '_save_next' in request.POST
            for orderitem in self.object.orderitem_set.all():
                orderitem.delete()
            
            formset.save()
            form = self.get_form()
            if form.fields and form.is_valid():
                if form.cleaned_data.get('do_photo'):
                    orgservice_photo = OrgService.objects.get(
                        org=self.object.loru,
                        service__name=Service.SERVICE_PHOTO,
                    )
                    price_photo = decimal.Decimal(
                        form.cleaned_data.get('price_photo', '0')
                    )
                    ServiceItem.objects.get_or_create(
                        order=self.object,
                        orgservice=orgservice_photo,
                        cost=price_photo,
                    )
                else:
                    self.object.serviceitem_set. \
                        filter(orgservice__service__name=Service.SERVICE_PHOTO).delete()

            write_log(self.request, self.object, _('Заказ сохранен'))
            if go_next:
                return redirect('order_burial', self.object.pk)
            else:
                msg = _("<a href='%(order_edit)s'>Заказ %(pk)s</a> сохранен") % dict(
                    order_edit=reverse('order_edit', args=[self.object.pk]),
                    pk=self.object.pk,
                )
                messages.success(self.request, msg)
                return redirect('.')
        else:
            messages.error(self.request, _("Обнаружены ошибки"))
            return self.get(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        self.request = request
        try:
            return render(request, self.template_name, self.get_context_data())
        except Order.DoesNotExist:
            raise Http404

order_products = OrderEditProducts.as_view()

class OrderInfo(LORURequiredMixin, DetailView):
    template_name = 'order_info.html'

    def get_queryset(self):
        return Order.objects.filter(loru=self.request.user.profile.org, type=Order.TYPE_BURIAL)

    def get_context_data(self, **kwargs):
        data = super(OrderInfo, self).get_context_data(**kwargs)
        results = ResultFile.objects.filter(order=self.object). \
                  order_by('-is_title', '-date_of_creation')
        data.update({
            'comment_form': CommentForm(),
            'results': results,
        })
        return data

order_info = OrderInfo.as_view()

class OrderEditServices(OrderEditProducts):
    template_name = 'order_edit_service.html'

    def get_catafalque_form(self):
        return CatafalqueForm(data=self.request.POST or None, instance=self.get_object().get_catafalquedata())

    def get_add_info_form(self):
        return AddInfoForm(data=self.request.POST or None, instance=self.get_object().get_addinfodata())

    def get_coffin_form(self):
        return CoffinForm(data=self.request.POST or None, instance=self.get_object().get_coffindata())

    def get_context_data(self, **kwargs):
        data = {'order': self.get_object()}
        if self.get_object().has_catafalque():
            data.update({'catafalque_form': self.get_catafalque_form()})
        if self.get_object().has_catafalque() or self.get_object().has_coffin():
            data.update({'add_info_form': self.get_add_info_form()})
        if self.get_object().has_coffin():
            data.update({'coffin_form': self.get_coffin_form()})
        return data

    def post(self, request, *args, **kwargs):
        self.request = request
        self.catafalque_form = self.get_catafalque_form()
        self.add_info_form = self.get_add_info_form()
        self.coffin_form = self.get_coffin_form()
        catafalque_ok = not self.get_object().has_catafalque() or self.catafalque_form.is_valid()
        add_info_ok = not (self.get_object().has_coffin() or self.get_object().has_catafalque()) or self.add_info_form.is_valid()
        coffin_ok = not self.get_object().has_coffin() or self.coffin_form.is_valid()
        if catafalque_ok and add_info_ok and coffin_ok:
            self.object = self.get_object()

            if self.catafalque_form.is_valid():
                cat = self.catafalque_form.save(commit=False)
                cat.order = self.object
                cat.save()

            if self.add_info_form.is_valid():
                add_info = self.add_info_form.save(commit=False)
                add_info.order = self.object
                add_info.save()

            if self.coffin_form.is_valid():
                coffin = self.coffin_form.save(commit=False)
                coffin.order = self.object
                coffin.save()

            write_log(self.request, self.object, _('Заказ сохранен'))
            if '_save_next' in request.POST:
                return redirect('order_info', self.object.pk)
            else:
                msg = _("<a href='%(order_edit)s'>Заказ %(pk)s</a> сохранен") % dict(
                    order_edit=reverse('order_edit', args=[self.object.pk]),
                    pk=self.object.pk,
                )
                messages.success(self.request, msg)
                return redirect('.')
        else:
            messages.error(self.request, _("Обнаружены ошибки"))
            return self.get(request, *args, **kwargs)

order_services = OrderEditServices.as_view()

class PrintOrderView(LORURequiredMixin, DetailView):
    context_object_name = 'order'

    def get_queryset(self):
        return Order.objects.filter(loru=self.request.user.profile.org).distinct()

    @transaction.atomic
    def render_to_response(self, context, **response_kwargs):
        order = self.get_object()
        customerprofile = password = responsible = deadman = None
        # Приглашение в ХРАМ, если необходимо:
        #   - есть сервис в созданном заказе
        #   - есть захоронение, в котором ответственный с телефоном
        burial = order.burial
        if order.serviceitem_set.exists() and \
          burial and \
          burial.deadman and \
          burial.deadman.last_name and \
          burial.cemetery:
            responsible = burial.responsible or \
                          burial.place and burial.place.responsible or \
                          None
        if responsible and responsible.login_phone:
            try:
                customerprofile = CustomerProfile.objects.get(
                    login_phone=responsible.login_phone
                )
            except CustomerProfile.DoesNotExist:
                user, password = CustomerProfile.create_cabinet(responsible, self.request)
                customerprofile = user.customerprofile
            deadman = order.burial.deadman
            person = deadman.baseperson_ptr
            if not CustomPerson.objects.filter(person=person):
                addr = _('Кладбище: %s') % burial.cemetery.name
                if burial.area:
                    addr = _("%(addr)s, участок: %(area)s") % dict(addr=addr, area=burial.area.name)
                if burial.row:
                    addr = _("%(addr)s, ряд: %(row)s") % dict(addr=addr, row=burial.row)
                if burial.place_number:
                    addr = _("%(addr)s, место: %(place)s") % dict(addr=addr, place=burial.place_number)
                location = Location.objects.create(addr_str=addr)
                customplace = CustomPlace.objects.create(
                    user=customerprofile.user,
                    place=None,
                    name=addr,
                    address=location,
                    favorite_performer=self.request.user.profile.org,
                )
                order.customplace = customplace
                order.save()
                customperson = CustomPerson.objects.create(
                    customplace=customplace,
                    person=person,
                    user=customerprofile.user,
                    last_name=deadman.last_name,
                    first_name=deadman.first_name,
                    middle_name=deadman.middle_name,
                    is_dead=True,
                    birth_date=deadman.birth_date,
                    death_date=deadman.death_date,
                )
        context.update(dict(
            user=self.request.user,
            now=datetime.datetime.now(),
            customerprofile=customerprofile,
            deadman=deadman,
            password=password,
            print_now=True,
        ))
        report = make_report(
            user=self.request.user,
            msg=_("Счет-заказ"),
            obj=order,
            template='reports/order_yalta.html' \
                if self.request.user.profile.org.inn == '9103078189' \
                else 'reports/order.html',
            context=context,
        )
        return redirect('report_view', report.pk)

order_print = PrintOrderView.as_view()

class PrintOrderReceiptView(LORURequiredMixin, DetailView):
    context_object_name = 'order'

    def get_queryset(self):
        return Order.objects.filter(loru=self.request.user.profile.org).distinct()

    def render_to_response(self, context, **response_kwargs):
        order = self.get_object()
        context.update(dict(
            now=datetime.datetime.now(),
            user=self.request.user,
        ))
        report = make_report(
            user=self.request.user,
            msg=_("Квитанция покупателя"),
            obj=order,
            template='reports/order_receipt.html',
            context=context,
        )
        return redirect('report_view', report.pk)

order_receipt_print = PrintOrderReceiptView.as_view()

class PrintContractView(LORURequiredMixin, DetailView):
    context_object_name = 'order'

    def get_queryset(self):
        return Order.objects.filter(loru=self.request.user.profile.org).distinct()

    def render_to_response(self, context, **response_kwargs):
        context['now'] = datetime.datetime.now()
        context['user'] = self.request.user
        report = make_report(
            user=self.request.user,
            msg=_("Договор"),
            obj=self.get_object(),
            template='reports/contract.html',
            context=context,
        )
        return redirect('report_view', report.pk)

order_contract = PrintContractView.as_view()

class CommentView(LORURequiredMixin, DetailView):
    def get_queryset(self):
        return Order.objects.filter(loru=self.request.user.profile.org).distinct()

    def post(self, request, *args, **kwargs):
        write_log(request, self.get_object(), _('Комментарий: %s') % request.POST.get('comment'))
        return redirect('order_edit', self.get_object().pk)

order_comment = CommentView.as_view()

class AnnulateOrder(LORURequiredMixin, DetailView):
    def get_queryset(self):
        return Order.objects.filter(loru=self.request.user.profile.org).distinct()

    def post(self, request, *args, **kwargs):
        referer = kwargs.get('referer')
        http_referer = request.META.get('HTTP_REFERER', '')
        o = self.get_object()
        if request.GET.get('recover'):
            o.recover()
            messages.success(self.request, _('Заказ восстановлен'))
            write_log(request, o, _('Заказ восстановлен'))
        else:
            b = o.burial
            old_annulated = b.annulated if b else None
            o.annulate()
            messages.success(self.request, _('Заказ аннулирован'))
            write_log(request, o, _('Заказ аннулирован'))
            if b and b.annulated and not old_annulated:
                write_log(request, b, _('Захоронение аннулировано'))
        if referer == 'edit':
            return redirect('order_edit', o.pk)
        else:
            if http_referer:
                return redirect(http_referer)
            else:
                return redirect('order_list')

order_annulate = AnnulateOrder.as_view()

class ApiOrderStatus(APIView):

    def post(self, request, what, pk):
        """
        Установить статус заказа в what (advanced|paid)
        """
        if what.lower() not in (Order.STATUS_ADVANCED, Order.STATUS_PAID,):
            raise Http404
        o = get_object_or_404(Order, loru=request.user.profile.org, pk=pk)
        if what == Order.STATUS_ADVANCED and not o.is_paid():
            if o.is_advanced():
                o.status = Order.STATUS_POSTED
                msg = _('Заказ: отмена получения аванса')
            else:
                o.status = Order.STATUS_ADVANCED
                msg = _('Заказ: получен аванс')
            o.save()
            write_log(request, o, msg)
        elif what == Order.STATUS_PAID:
            if o.is_paid():
                o.status = Order.STATUS_POSTED
                msg = _('Заказ: отмена получения оплаты')
            else:
                o.status = Order.STATUS_PAID
                msg = _('Заказ: оплачен')
            o.save()
            write_log(request, o, msg)
        return Response(data={}, status=200)

api_order_status = ApiOrderStatus.as_view()

class OrderBurialView(LORURequiredMixin, RequestToFormMixin, UpdateView):
    """
    Cоздание или привязка захоронения к заказу
    """
    template_name = 'order_burial.html'
    form_class = OrderBurialForm

    def get_queryset(self):
        return Order.objects.filter(loru=self.request.user.profile.org, type=Order.TYPE_BURIAL)

    def get(self, request, *args, **kwargs):
        order = self.get_object()
        burial = order.burial
        if burial:
            if burial.is_edit() and not burial.annulated:
                return redirect(reverse('edit_burial', args=[burial.pk]) + '?order=%s' % order.pk)
            else:
                return redirect(reverse('view_burial', args=[burial.pk]) + '?order=%s' % order.pk)
        return super(OrderBurialView, self).get(request, *args, **kwargs)
        
    def form_valid(self, form):
        if self.object.burial:
            # - форма привязала к этому заказу захоронение
            write_log(self.request, self.object.burial, _('Захоронение прикреплено к заказу %s') % self.object.pk)
            write_log(self.request, self.object, _('Заказ: прикреплено захоронение %s') % self.object.burial.pk)
            msg = _("<a href='%(order_edit)s'>Заказ %(pk)s</a>: прикреплено захоронение") % dict(
                order_edit=reverse('order_edit', args=[self.object.pk]),
                pk=self.object.pk,
            )
            messages.success(self.request, msg)
            return redirect('.')
        # - форма отдала "Создать новое захоронение"
        return redirect(reverse('create_burial') + '?order=%s' % self.object.pk)

order_burial = OrderBurialView.as_view()

class ProductCategories(APIView):

    def get(self, request):
        loru_ids = request.GET.getlist('filter[supplier]')
        while loru_ids.count(''):
            loru_ids.remove('')
        if loru_ids:
            qs = Q(product__loru__pk__in=loru_ids)
        else:
            qs = Q()
        onlyOpt = request.GET.get('filter[onlyOpt]')
        if onlyOpt and onlyOpt == 'true':
            qs &= Q(product__is_wholesale=True)
        categories = [ProductCategorySerializer(pc).data \
            for pc in ProductCategory.objects.filter(qs).order_by('name').distinct()
        ]
        return Response(
            status=200,
            data = {
                # Раньше это был viewset с пажинацией и на то настроен front-end
                #
                'count': len(categories),
                'next': None,
                'previous': None,
                'results': categories,
        })

api_catalog_categories = ProductCategories.as_view()

class ProductsViewSet(ProductCategoryQsMixin, viewsets.ReadOnlyModelViewSet):
    """
    Показ продуктов в публичном каталоге только!!!
    """
    model = Product
    serializer_class = ProductsSerializer

    def get_queryset(self):
        qs = Q(is_archived=False)

        store_ids = self.request.GET.getlist('filter[supplierStore]')
        while store_ids.count(''):
            store_ids.remove('')
        if store_ids:
            qs &= Q(loru__store__pk__in=store_ids)

        loru_ids = self.request.GET.getlist('filter[supplier]')
        while loru_ids.count(''):
            loru_ids.remove('')
        if loru_ids:
            qs &= Q(loru__pk__in=loru_ids)

        if self.request.GET.get('filter[price_from]'):
            qs &= Q(price__gte=self.request.GET.get('filter[price_from]'))
        if self.request.GET.get('filter[price_to]'):
            qs &= Q(price__lte=self.request.GET.get('filter[price_to]'))

        qs &= self.product_category_qs()

        q_public_whole = Q(is_public_catalog=True)
        if self.request.GET.get('filter[productType]', '').lower() == 'opt':
            q_public_whole = Q(is_wholesale=True)
        qs &= q_public_whole

        available_for_visit = str_to_bool_or_None(self.request.GET.get('filter[isAvailableForVisitOrder]'))
        if available_for_visit:
            qs &= Q(productcategory__pk__in=ProductCategory.AVAILABLE_FOR_VISIT_PKS)

        ordered = None
        orders = {'price': 'price', 'date': 'dt_created', }
        directions = {'asc': '', 'desc': '-', }
        for order in orders:
            direction = self.request.GET.get('order[%s]' % order)
            if direction and direction in directions:
                ordered = directions[direction] + orders[order]
                break
        
        filter = Product.objects.filter(qs)
        if ordered:
            filter = filter.order_by(ordered)
        filter = filter.distinct()

        offset = self.request.GET.get('offset') and int(self.request.GET.get('offset'))
        limit = self.request.GET.get('limit') and int(self.request.GET.get('limit'))

        if offset and limit:
            filter = filter[offset:offset+limit]
        elif offset:
            filter = filter[offset:]
        elif limit:
            filter = filter[:limit]
        
        return filter

class ProductsOptViewSet(ProductCategoryQsMixin, viewsets.ReadOnlyModelViewSet):
    """
    Показ продуктов оптовика-поставщика

    api/optplaces/suppliers/(?P<loru_pk>\d+)/products
    """
    model = Product

    def get_queryset(self, *args, **kwargs):
        qs = Q(
            loru__pk=self.kwargs['loru_pk'],
            is_wholesale=True,
        )
        qs &= self.product_category_qs()
        return Product.objects.filter(qs)

    def get_serializer(self, *args, **kwargs):
        try:
            is_wholesale_with_vat = Org.objects.get(pk=self.kwargs['loru_pk']).is_wholesale_with_vat
        except Org.DoesNotExist:
            is_wholesale_with_vat = False
        return ProductsOptSerializer(
            self.get_queryset(),
            context = dict(
                request=self.request,
                view=self,
                is_wholesale_with_vat=is_wholesale_with_vat,
            ),
            many=True,
        )

class ProductInfoView(APIView):

    def get(self, request, product_slug):
        product = get_object_or_404(Product, slug=product_slug)
        show_wholesale = is_trade_user(request.user) or is_supervisor(request.user)
        return Response(
            status=200,
            data=ProductInfoSerializer(product, context=dict(
                request=request,
                show_wholesale=show_wholesale,
            )).data,
        )

api_catalog_products_detail = ProductInfoView.as_view()

class ApiLoruProductPlaces(APIView):
    """
    Обновление статусов продуктов на площадках (ОМС)

    Пример:
    [
        {
            "id": 1,
            "places": [
            {
                "id": 5, 
                "status": "disable"
            },
            {
                "id": 8, 
                "status": "up"
            }
            ]
        },
        {
            "id": 2,
            "places": [
            {
                "id": 5, 
                "status": "enable"
            }
            ]
        }
    ]
    """
    
    permission_classes = (PermitIfTrade,)

    @transaction.atomic
    def post(self, request, format=None):
        # Соответствие входных данных с константами в поле Rate.action
        rate_action = {
            'disable': Rate.RATE_ACTION_DISABLE,
            'enable': Rate.RATE_ACTION_PUBLISH,
            'up': Rate.RATE_ACTION_UPDATE,
        }
        data = []
        catalog_org_pk = Org.get_catalog_org_pk()
        for p in request.data:
            try:
                product = Product.objects.get(pk=p['id'], loru=request.user.profile.org)
                data_p = { 'id': p['id'], 'places': [catalog_org_pk] }
                for o in p['places']:
                    if o['id'] == catalog_org_pk:
                        try:
                            if rate_action[o['status']] == Rate.RATE_ACTION_DISABLE:
                                product.is_public_catalog = False
                            elif rate_action[o['status']] in \
                                    (Rate.RATE_ACTION_PUBLISH, Rate.RATE_ACTION_UPDATE,):
                                product.is_public_catalog = True
                            product.save()
                            write_log(
                                request,
                                product,
                                _("Добавлен в публичный каталог") if product.is_public_catalog else _("Изъят из публичного каталога"),
                            )
                            data.append(data_p)
                        except IndexError:
                            pass
            except Product.DoesNotExist:
                pass
        return Response(data=data, status=200)

api_loru_product_places = ApiLoruProductPlaces.as_view()

class UghPublishedProductsViewSet(viewsets.ViewSet):
    queryset = Product.objects.none()
    permission_classes = (PermitIfTrade,)
    
    def list(self, request):
        data=[]
        catalog_org_pk = Org.get_catalog_org_pk()
        for p in Product.objects.filter(loru=request.user.profile.org, is_archived=False).order_by('pk'):
            data_p = {
                'id': p.pk,
                'name': p.name,
                'availableOnPlaces': [],
                'category': {
                    'id': p.productcategory.pk,
                    'name': p.productcategory.name,
                },
            }
            if p.is_public_catalog:
                data_p['availableOnPlaces'] = [catalog_org_pk,]
            data.append(data_p)
        return Response(status=200, data=data)

class OptOrderMixin(APIView):

    def put_item(self, order, product, count, comment):
        """
        Забить позицию интернет-заказа продуктом product
        """
        return OrderItem.objects.create(
            order=order,
            product=product,
            quantity=decimal.Decimal(count),
            comment=comment or '',
            cost=product.price_wholesale,
            name=product.name,
            measure=product.measure,
            description=product.description,
            productcategory=product.productcategory,
            productcategory_name=product.productcategory.name,
            productgroup=product.productgroup,
            productgroup_name=product.productgroup.name if product.productgroup else '',
            productgroup_description=product.productgroup.description if product.productgroup else '',
            is_wholesale_with_vat=order.loru.is_wholesale_with_vat,
        )

    def email_notifications(self, order, is_new_opt_order, new_status=False):
        """
        """
        email_from = settings.DEFAULT_FROM_EMAIL
        number_verbose = order.number_verbose()
        if order.applicant_organization and order.applicant_organization.email:
            email_to = (order.applicant_organization.email, )
            email_subject = "%s: %s %s" % (
                _("Похоронное Дело"),
                _("создан заказ") if is_new_opt_order else _("изменен заказ"),
                number_verbose,
            )
            email_text = render_to_string(
                            'opt_order_notification.txt',
                            {
                                'preambule': _("Создан") if is_new_opt_order else _("Изменен"),
                                'front_end_url': get_front_end_url(self.request),
                                'order': order,
                                'to_customer': True,
                                'new_status': new_status,
                            }
            )
            EmailMessage(email_subject, email_text, email_from, email_to,).send()
        if order.loru.email:
            email_to = (order.loru.email, )
            email_subject = "%s: %s %s" % (
                _("Похоронное Дело"),
                _("поступил заказ") if is_new_opt_order else _("изменен заказ"),
                number_verbose,
            )
            email_text = render_to_string(
                            'opt_order_notification.txt',
                            {
                                'preambule': _("Поступил") if is_new_opt_order else _("Изменен"),
                                'front_end_url': get_front_end_url(self.request),
                                'order': order,
                                'to_customer': False,
                                'new_status': new_status,
                            }
            )
            EmailMessage(email_subject, email_text, email_from, email_to,).send()
        if not settings.DEBUG:
            # Отправка смс поставщику
            if order.loru.sms_phone:
                supplier_email = " (email: %s)" % order.loru.email if order.loru.email else ""
                text =  _("%s zakaz N %s summa %s") % (
                    get_front_end_url(self.request).rstrip('/'),
                    number_verbose,
                    order.total_float(),
                )
                if is_new_opt_order:
                    email_error_text = "Поставщик %s%s не получил СМС- уведомление о новом заказе" % \
                                        (order.loru.name, supplier_email,)
                else:
                    email_error_text = _("Поставщик %s%s не получил СМС- уведомление об изменении заказа %s") % \
                                        (order.loru.name, supplier_email, number_verbose, )
                send_sms(
                    phone_number=order.loru.sms_phone,
                    text=text,
                    email_error_text=email_error_text,
                )
            elif order.loru.email:
                EmailMessage(
                    subject=_('Похоронное Дело: телефон смс- уведомений'),
                    body=_(
                        'Невозможно доставить СМС %s заказе № %s.\n'
                        '\n'
                        'В свойствах вашей организации: %s\n'
                        'не указан телефон для СМС- уведомлений.\n'
                        '\n'
                        'Это можно исправить: %s\n'
                    ) % (
                        _('о новом') if is_new_opt_order else _('об измененном'),
                        number_verbose,
                        order.loru.name,
                        get_host_url(self.request) + reverse('edit_org', args=(order.loru.pk,)).lstrip('/'),
                    ),
                    from_email=email_from,
                    to=(order.loru.email,),
                ).send(fail_silently=True)

class ApiOptPlacesOrders(OptOrderMixin, APIView):
    """
    Интернет-заказ товаров

    Пример входных данных:
    {
        "products": [
            {
            "id": 62,
            "count": 2
            },
            {
            "id": 57,
            "count": 1
            }
        ],
        "comment": "Текст комментария",
        "customer": {
            "id": 2,
            # или, если заказал по телефону
            "title": "Директор Рога и Копыта",
            "phoneNumber": "375291234567",
            "address": "Одесса Малая Арнаутская"
        }
    }
    Заказывает покупатель у поставщика.
    Покупатель может быть указан в customer, или если не указан, то это выполнивший
    логин.
    Поставщик - в id продуктов. Эти id проверяются на то, чтоб им соответствовали
    продукты, и чтоб они относились к одному поставщику. Если проверка не пройдет,
    то возвращается status=400, message=”что произошло”. Если проверка успешна,
    то формируется новый заказ, статус код - 200 
    """
    permission_classes = (PermitIfTrade,)

    @transaction.atomic
    def post(self, request):
        status_code=200
        data = dict(status='success')
        year = datetime.datetime.now().year
        try:
            customer= title = phones = address = None
            customer_input = request.data.get("customer")
            if customer_input:
                if customer_input.get('id'):
                    try:
                        customer = Org.objects.get(pk=customer_input['id'])
                    except Org.DoesNotExist:
                        raise ServiceException(
                            _("Покупатель с сustomer['id']==%s отсутствует") % customer_input['id']
                        )
                elif set(('title', 'phoneNumber', 'address',)).intersection(list(customer_input.keys())):
                    title = customer_input.get('title', '')
                    phones = customer_input.get('phoneNumber')
                    addr_str = customer_input.get('address')
                    if addr_str:
                        address = Location.objects.create(addr_str=addr_str)
                else:
                    raise ServiceException(_('Невозможно определить покупателя'))
            else:
                customer = request.user.profile.org
            supplier = None
            products = request.data.get("products",[])
            comment = request.data.get("comment",'')
            for p in products:
                try:
                    product = Product.objects.get(pk=p['id'])
                    if supplier is None:
                        supplier = product.loru
                        order = Order.objects.create(
                            type=Order.TYPE_TRADE,
                            loru=supplier,
                            applicant_organization=customer,
                            status=Order.STATUS_POSTED,
                            payment=Order.PAYMENT_WIRE,
                            dt=datetime.date.today(),
                            title=title or '',
                            phones=phones,
                            address=address,
                        )
                        if comment:
                            OrderComment.objects.create(
                                order=order,
                                user=request.user,
                                comment=comment,
                            )
                    else:
                        if product.loru != supplier:
                            raise ServiceException(_('В списке товаров таковые от разных поставщиков'))
                    self.put_item(order, product, p['count'], p.get('comment'))
                    data['id'] = order.pk
                except Product.DoesNotExist:
                    raise ServiceException(_('Не найден товар/услуга Id=%s') % p['id'])
        except ServiceException as excpt:
            transaction.set_rollback(True)
            status_code=400
            data['status'] = 'error'
            data['message'] = excpt.args[0]
        else:
            self.email_notifications(order, is_new_opt_order=True)
        return Response(data=data, status=status_code)

    def get(self, request):
        org = request.user.profile.org
        qs = Q(loru=org) | Q(applicant_organization=org)
        opt_orders = Order.objects.filter(qs & Q(type=Order.TYPE_TRADE)).order_by('-dt_created').distinct()
        return Response(
            status=200,
            data=OptOrdersSerializer(opt_orders, many=True, context=dict(
                request=request,
            )).data,
        )

api_optplaces_orders = ApiOptPlacesOrders.as_view()

class OptOrderderInfoView(OptOrderMixin, APIView):
    permission_classes = (PermitIfTrade,)

    def instance_permitted(self, request, pk):
        """
        Просматривать и править заказ может лишь поставщик или покупатель, если имеется
        """
        order = get_object_or_404(Order, pk=pk, type=Order.TYPE_TRADE)
        org = request.user.profile.org
        if order.loru == org or order.applicant_organization and order.applicant_organization == org:
            return order, None
        else:
            return None, Response(
                status=403,
                data={"detail": "Either it is not a trade order, or not authorized: you are not customer nor supplier."},
            )

    def get(self, request, pk):
        order, response = self.instance_permitted(request, pk)
        if response:
            return response
        return Response(
            status=200,
            data=OptOrderInfoSerializer(order, context=dict(
                request=request,
            )).data,
        )

    @transaction.atomic
    def put(self, request, pk):
        order, response = self.instance_permitted(request, pk)
        if response:
            return response
        status_code=200
        data = dict(status='success')
        products = request.data.get("products")
        if products is None:
            raise ServiceException(_('Не задан список товаров, пусть даже пустой'))
        try:
            OrderItem.objects.filter(order=order).delete()
            for p in products:
                try:
                    product = Product.objects.get(pk=p['id'])
                    if product.loru != order.loru:
                        raise ServiceException(_('Товара Id=%s нет среди товаров поставщика заказа') % p['id'])
                    if p.get('count'):
                        self.put_item(order, product, p['count'], p.get('comment'))
                except Product.DoesNotExist:
                    raise ServiceException(_('Не найден товар/услуга Id=%s') % p['id'])
            comment = request.data.get("comment")
            if comment is not None:
                try:
                    ordercomment = OrderComment.objects.filter(order=order)[0]
                except IndexError:
                    OrderComment.objects.create(
                        order=order,
                        user=request.user,
                        comment=comment,
                    )
                else:
                    ordercomment.comment = comment
                    ordercomment.save()
            # В любом случае сохранить, чтоб подправилась dt_modified
            order.save()
        except ServiceException as excpt:
            transaction.set_rollback(True)
            status_code=400
            data['status'] = 'error'
            data['message'] = excpt.args[0]
        else:
            self.email_notifications(order, is_new_opt_order=False)
        return Response(data=data, status=status_code)

api_optplaces_orders_detail = OptOrderderInfoView.as_view()

class ApiLoruProductTypesView(APIView):
    permission_classes = (PermitIfTrade,)

    def get(self, request):
        return Response(
            data=[ dict(
                    id=pt[0],
                    name =str(pt[1]),
                   )
                    for pt in Product.PRODUCT_TYPES
            ],
            status=200,
        )

api_loru_product_types = ApiLoruProductTypesView.as_view()

class ApiProductList(ProductCategoryQsMixin, APIView):
    permission_classes = (PermitIfTrade,)
    parser_classes = (MultiPartParser,)

    def get(self, request):
        qs = Q(loru=request.user.profile.org)
        qs &= self.product_category_qs()
        is_archived = request.GET.get('filter[is_archived]')
        if is_archived:
            try:
                is_archived = bool(int(is_archived))
            except ValueError:
                is_archived = None
        # По умолчанию показываем не-архивные продукты
        if is_archived is None:
            qs &= Q(is_archived=False)
        else:
            qs &= Q(is_archived=is_archived)

        data = [ ProductEditSerializer(p, context=dict(request=request)).data \
                for p in Product.objects.filter(qs)
        ]
        return Response(data=data, status=200)

    @transaction.atomic
    def post(self, request):
        try:
            required_not_got = list()
            for f in (
                'name',
                'description',
                'categoryId',
                    ):
                if not request.data.get(f):
                    required_not_got.append(f)
            if required_not_got:
                raise ServiceException(_("Не заданы параметры: %s") % ", ".join(required_not_got))
            serializer = ProductEditSerializer(data=request.data, context={ 'request': request, })
            if serializer.is_valid():
                product = serializer.save()
                if not product.sku or not product.sku.strip():
                    product.sku = product.pk
                    product.save()
                    serializer = ProductEditSerializer(product, context={ 'request': request, })
                return Response(serializer.data, status=200)
            return Response(serializer.errors, status=400)
        except ServiceException as excpt:
            return Response(data={'status': 'error', 'message': excpt.args[0]}, status=400)

api_product_list = ApiProductList.as_view()

class ApiProductDetail(APIView):
    permission_classes = (PermitIfTrade,)
    parser_classes = (MultiPartParser,)

    def get_object(self, request, pk):
        try:
            return Product.objects.get(loru=request.user.profile.org, pk=pk)
        except Product.DoesNotExist:
            raise Http404

    def get(self, request, pk):
        product = self.get_object(request, pk)
        serializer = ProductEditSerializer(product, context=dict(request=request))
        return Response(serializer.data)

    @transaction.atomic
    def put(self, request, pk):
        product = self.get_object(request, pk)
        serializer = ProductEditSerializer(
            product,
            data=request.data,
            context=dict(request=request),
        )
        if serializer.is_valid():
            try:
                serializer.save()
                return Response(serializer.data, status=200)
            except ServiceException as excpt:
                return Response(data={'status': 'error', 'message': excpt.args[0]}, status=400)
        return Response(serializer.errors, status=400)

api_product_detail = ApiProductDetail.as_view()

class ApiServicesView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        serializer = ServiceSerializer(Service.objects.all(), many=True)
        return Response(serializer.data, status=200)

api_services = ApiServicesView.as_view()

class ApiOrgServicesMixin(object):
    
    class Data(object):
        service=None
        measures = []
        orgservice = None
        measures_get = []

    def check_org_id(self, request, org_id):
        org, message = request.user.profile.org, ''
        if str(org.pk) != str(org_id):
            message = _("Org_id %s не соответствует организации, выполняющей запрос") % org_id
        return org, message
    
    def check_input_message(self, request, org_id, service_name=None):
        self.data = self.Data()
        self.data.measures_get = request.data.get('measures', [])

        org, message = self.check_org_id(request, org_id)
        if message:
            return message

        if not service_name:
            service_name = request.data.get('type')
        try:
            self.data.service = Service.objects.get(name=service_name)
        except Service.DoesNotExist:
            return _("Сервис %s неизвестен в системе") % service_name

        try:
            self.data.orgservice = OrgService.objects.get(org=org, service=self.data.service)
        except OrgService.DoesNotExist:
            if request.method == 'POST' and not self.data.measures_get:
                return _("У организации нет цен по сервису %s. Нельзя его активировать, не указав цены") % service_name
            # - put (задать цены):          сервис будет создан у организации
            # - delete (де-активировать):   нет сервиса, не будет и после запроса,
            #                               т.е. сервис останется неактивированным у организации

        if request.method == 'PUT' and not self.data.measures_get:
            return _("Изменение цен по сервису. Не указаны цены")

        self.data.measures = Measure.objects.filter(service=self.data.service)
        measure_names = [v['name'] for v in self.data.measures.values('name')]
        for m in self.data.measures_get:
            if m['name'] not in measure_names:
                return _("Нет единицы измерения %s у услуги %s") % (m['name'], service_name)
        return ''

    def put_prices(self, request):
        if self.data.orgservice:
            orgservice = self.data.orgservice
            if request.method == 'POST' and not orgservice.enabled:
                orgservice.enabled = True
                orgservice.save()
        else:
            orgservice = OrgService.objects.create(
                org=request.user.profile.org,
                service=self.data.service,
                enabled=request.method == "POST",
            )
            self.data.orgservice = orgservice
        for measure in self.data.measures:
            price = None
            for m in self.data.measures_get:
                if m['name'] == measure.name:
                    price = m['price']
                    break
            orgserviceprice, created_ = OrgServicePrice.objects.get_or_create(
                orgservice=orgservice,
                measure=measure,
                defaults=dict(price=price or 0.00)
            )
            if not created_ and price is not None:
                orgserviceprice.price = price
                orgserviceprice.save()

class ApiOrgServicesView(ApiOrgServicesMixin, APIView):
    permission_classes = (PermitIfTrade,)

    def get(self, request, org_id):
        org, message = self.check_org_id(request, org_id)
        if message:
            return Response(data=dict(status='error', message=message), status=200)
        return Response(
            data=OrgServiceSerializer(OrgService.objects.filter(org=org), many=True).data,
            status=200
        )

    @transaction.atomic
    def post(self, request, org_id):
        """
        Активизировать сервис у организации org_id
        
        Вх.данные, пример:
        {
        "type": "photo",
        "measures": [
          {
          "name": "unit",
          "value": 234
          }
         ]
        }
        """
        message = self.check_input_message(request, org_id)
        if message:
            return Response(data=dict(status='error', message=message), status=400)
        self.put_prices(request)
        return Response(data=dict(status='success'), status=200)

api_org_services = ApiOrgServicesView.as_view()

class ApiOrgServicesEditView(ApiOrgServicesMixin, APIView):
    permission_classes = (PermitIfTrade,)

    @transaction.atomic
    def put(self, request, org_id, service_name):
        """
        Править сервис service_name у организации org_id
        
        Вх.данные, пример:
        {
        "measures": [
          {
          "name": "unit",
          "value": 234
          }
         ]
        }
        """
        message = self.check_input_message(request, org_id, service_name)
        if message:
            return Response(data=dict(status='error', message=message), status=400)
        self.put_prices(request)
        return Response(data=dict(status='success'), status=200)

    def delete(self, request, org_id, service_name):
        """
        Де-активация сервиса service_name у организации org_id
        """
        message = self.check_input_message(request, org_id, service_name)
        if message:
            return Response(data=dict(status='error', message=message), status=400)
        orgservice = self.data.orgservice
        if orgservice and orgservice.enabled:
            orgservice.enabled = False
            orgservice.save()
        return Response(data=dict(status='success'), status=200)

api_org_services_edit = ApiOrgServicesEditView.as_view()

class ApiServicePriceMixin(object):

    class Data(object):
        service = None
        service_name = None
        orgservice=None
        orgservice_delivery = None
        customplace = None
        latitude = None
        longitude = None
        need_delivery = False
        org = None

    def check_input_message(self, request):
        self.data = self.Data()
        try:
            if request.method == 'GET':
                req_dict = request.GET
            elif request.method == 'POST':
                req_dict = request.data

            self.data.service_name = req_dict.get('type')
            try:
                self.data.service = self.data.service_name and Service.objects.get(name=self.data.service_name)
            except Service.DoesNotExist:
                self.data.service = None
            if not self.data.service:
                raise ServiceException(_('Сервис не задан или неизвестен'))

            if request.method == 'POST':
                org = req_dict.get('performerId')
                try:
                    org = org and Org.objects.get(pk=org)
                except Org.DoesNotExist:
                    pass
                if not org:
                    raise ServiceException(_("Id исполнителя не задан или не найден среди организаций"))
                if org.type not in (Org.PROFILE_LORU, ):
                    raise ServiceException(_("performerId %s - не ЛОРУ (поставщик услуг)") % org.pk)
                try:
                    self.data.orgservice = OrgService.objects.get(
                        org=org,
                        service=self.data.service,
                        enabled=True
                    )
                except OrgService.DoesNotExist:
                    raise ServiceException(_("Сервис %s не активирован у организации Id=%s") % (
                          self.data.service_name,
                          org.pk,
                    ))
                self.data.org = org

            customplace_id = req_dict.get('placeId')
            try:
                self.data.customplace = customplace_id and CustomPlace.objects.get(pk=customplace_id)
            except CustomPlace.DoesNotExist:
                self.data.customplace = None
            if not self.data.customplace:
                raise ServiceException(_('Не задан placeId или не найдено место'))
            if self.data.customplace.user != request.user:
                raise ServiceException(_('Пользователь %s не имеет прав на запрос по этому месту') % request.user.username)
            
            self.data.need_delivery = self.data.service_name in (Service.SERVICE_PHOTO, Service.SERVICE_DELIVERY, )
            if self.data.need_delivery:
                latitude = longitude = None
                if request.method == 'GET':
                    latitude = req_dict.get('location[latitude]')
                    longitude = req_dict.get('location[longitude]')
                if latitude is None or longitude is None:
                    latitude = self.data.customplace.address and self.data.customplace.address.gps_y
                    longitude = self.data.customplace.address and self.data.customplace.address.gps_x
                if (latitude is None or longitude is None) and self.data.customplace.place:
                    place_location = self.data.customplace.place.location()
                    if place_location:
                        latitude = place_location['latitude']
                        longitude = place_location['longitude']
                if latitude is None or longitude is None:
                    raise ServiceException(_('Не заданы или не удалось выяснить координаты места'))
                self.data.latitude = float(latitude)
                self.data.longitude = float(longitude)
                if request.method == 'POST' and self.data.service_name != Service.SERVICE_DELIVERY:
                    try:
                        self.data.orgservice_delivery = OrgService.objects.get(
                            org=org,
                            service__name=Service.SERVICE_DELIVERY,
                            enabled=True
                        )
                    except OrgService.DoesNotExist:
                        raise ServiceException(_('Услуга %s требует еще активной услуги доставки у организации') % (
                            self.data.service_name,
                        ))
                    

        except ServiceException as excpt:
            return excpt.args[0]
        return ''

    def get_price_service(self, org, service):
        """
        Цена за услугу у огранизации
        """
        price_service = 0
        if service.name == Service.SERVICE_DELIVERY:
            # цена за доставку считается не по организации, а по складам
            pass
        elif service.name in (Service.SERVICE_PHOTO, ):
            price_service = OrgServicePrice.objects.get(
                orgservice__service=service,
                orgservice__org=org,
                measure__name='unit',
            ).price
        return decimal.Decimal(round(float(price_service), org.currency.rounding))

    def get_price_delivery(self, org, km, kg=None, m3=None):
        """
        Цена за доставку XX kg и/или yy m3 груза на расстояние km в организации org
        """
        result = 0.00
        for m in OrgServicePrice.objects.filter(
            orgservice__org=org,
            orgservice__service__name=Service.SERVICE_DELIVERY
            ).values('measure__name', 'price'):
            if m['measure__name'] == 'km':
                result += float(m['price']) * km
            elif kg and m['measure__name'] == 'kg':
                result += float(m['price']) * kg * km
            elif m3 and m['measure__name'] == 'm3':
                result += float(m['price']) * m3 * km
        return decimal.Decimal(round(result, org.currency.rounding))

class ApiClientAvailablePerformersView(ApiServicePriceMixin, APIView):
    permission_classes = (PermitIfCabinet,)

    def get(self, request):
        """
        Получить список возможных исполнителей работы
        
        Выбираются исполнители, подписавшиеся на сервис.
            - если сервис требует транспортных расходов, то еще и подписавшиеся
              на сервис delivery.
        Из них выбираются те, кто 
            - имеют склады с коодинатами
            - имеют account в платежной системе, пока только webpay
        Из этих складов выбираются ближайший
        """
        message = self.check_input_message(request)
        if message:
            return Response(data=dict(status='error', message=message), status=400)

        q = Q(
            loru__orgservice__service=self.data.service,
            loru__orgservice__enabled=True,
        )
        q_payable = Q(loru__orgwebpay__org__isnull=False)
        q &= q_payable
        if self.data.need_delivery:
            q &= Q(
                address__gps_x__isnull=False,
                address__gps_y__isnull=False,
            )
            if self.data.service_name != Service.SERVICE_DELIVERY:
                orgs = [orgservice.org for orgservice in OrgService.objects.filter(
                    service=self.data.service,
                    enabled=True,
                )]
                q &= Q(loru__in=orgs)
            customer_loc = LatLon(Latitude(self.data.latitude), Longitude(self.data.longitude))

        data = []
        org = None
        # Больше длины экватора
        max_distance = 41000.0
        for store in Store.objects.filter(q).order_by('loru'):
            # Идем по складам одного поставщика, потом по складам другого поставщика...
            if org is None:
                org = store.loru
                # Цена за любую услугу, кроме delivery
                price_org = self.get_price_service(org, self.data.service)
                distance = max_distance
                location = None

            if store.loru != org:
                if self.data.need_delivery:
                    kwargs = dict(km=distance)
                    price_org += self.get_price_delivery(org, **kwargs)
                data.append(dict(
                    id=org.pk,
                    name=org.name,
                    domainName=org.subdomain,
                    location=location,
                    price=round(price_org, org.currency.rounding),
                    currency=org.currency.code,
                ))
                org = store.loru
                price_org = self.get_price_service(org, self.data.service)
                distance = max_distance
                location = None

            if self.data.need_delivery:
                store_loc = LatLon(Latitude(store.address.gps_y), Longitude(store.address.gps_x))
                cur_distance = store_loc.distance(customer_loc)
                if distance > cur_distance:
                    distance = cur_distance
                    location = store.address.location_dict()

        if org:
            if self.data.need_delivery:
                kwargs = dict(km=distance)
                price_org += self.get_price_delivery(org, **kwargs)
            data.append(dict(
                id=org.pk,
                name=org.name,
                domainName=org.subdomain,
                location=location,
                price=float(price_org),
                currency=org.currency.code,
            ))

        return Response(data=data, status=200)

api_client_available_performers = ApiClientAvailablePerformersView.as_view()

class ApiShopPlacesView(ApiServicePriceMixin, APIView):

    def get(self, request, org_pk, customplace_pk):
        """
        Получить сумму на выполнение заказа фото на место (CustomPlace!)

        Выполняется анонимным пользователем
        """
        response_data = dict()
        status_code = 200
        try:
            org = get_object_or_404(Org, pk=org_pk)
            customplace = get_object_or_404(CustomPlace, pk=customplace_pk)
            place_location = customplace.location()
            if not place_location:
                raise ServiceException(_("Место не имеет координат"))

            service_name = Service.SERVICE_PHOTO
            service = Service.objects.get(name=service_name)
            try:
                orgservice = OrgService.objects.get(
                    org=org,
                    service=service,
                    enabled=True
                )
            except OrgService.DoesNotExist:
                raise ServiceException(_("Сервис %s не активирован у организации Id=%s") % (
                        service_name,
                        org.pk,
                ))

            try:
                orgservice_delivery = OrgService.objects.get(
                    org=org,
                    service__name=Service.SERVICE_DELIVERY,
                    enabled=True
                )
            except OrgService.DoesNotExist:
                raise ServiceException(_('Услуга %s требует еще активной услуги доставки у организации') % (
                    service_name,
                ))

            store_qs = Store.objects.filter(
                loru=org,
                address__gps_x__isnull=False,
                address__gps_y__isnull=False,
            )
            if not store_qs.count():
                raise ServiceException(_('У организации нет складов с координатами'))
            place_latlon = LatLon(Latitude(place_location['latitude']), Longitude(place_location['longitude']))

            price_org = self.get_price_service(org, service)

            # Вычисляем расстояние от места до ближайшего склада
            # Больше длины экватора
            distance = 41000.0
            for store in store_qs:
                store_loc = LatLon(Latitude(store.address.gps_y), Longitude(store.address.gps_x))
                distance = min(distance, store_loc.distance(place_latlon))

            price_org += self.get_price_delivery(org=org, km=distance)
            response_data['base_price'] = float(price_org)
            response_data['services'] = [
                dict(title=service.title),
                dict(title=orgservice_delivery.service.title),
            ]

        except ServiceException as excpt:
            status_code = 400
            response_data['status'] = 'error'
            response_data['message'] = excpt.args[0]
        return Response(data=response_data, status=status_code)

api_shops_places = ApiShopPlacesView.as_view()

class ApiClientOrdersView(ApiServicePriceMixin, APIView):
    permission_classes = (PermitIfCabinet,)

    @transaction.atomic
    def post(self, request):
        """
        Принять предложение о создании заказа

        Вх. данные, пример:
        {
            "type": "photo",
            "performerId": 19,
            "placeId": 43, # CustomPlace, обязательно с координатами
            "comment": "Коментарий к заказу"
        }
        """
        message = self.check_input_message(request)
        if message:
            return Response(data=dict(status='error', message=message), status=400)

        distance = 41000.0
        price_org = self.get_price_service(self.data.org, self.data.service)
        closest_store = dict(lat=0, lng=0)
        if self.data.need_delivery:
            customer_loc = LatLon(Latitude(self.data.latitude), Longitude(self.data.longitude))
            for store in self.data.org.store_set.filter(
                address__gps_x__isnull=False,
                address__gps_y__isnull=False,
            ):
                store_loc = LatLon(Latitude(store.address.gps_y), Longitude(store.address.gps_x))
                distance_store = store_loc.distance(customer_loc)
                if distance_store < distance:
                    distance = distance_store
                    closest_store=dict(lat=store.address.gps_y, lng=store.address.gps_x)

            kwargs = dict(km=distance)
            price_delivery = self.get_price_delivery(self.data.org, **kwargs)
        else:
            price_delivery = 0.00

        login_phone = request.user.customerprofile.login_phone
        applicant = AlivePerson.objects.create(
            user=request.user,
            phones="+%s" % login_phone if login_phone else None,
            login_phone=login_phone,
            last_name=request.user.customerprofile.user_last_name,
            first_name=request.user.customerprofile.user_first_name,
            middle_name=request.user.customerprofile.user_middle_name,
        )
        order = Order(
            loru=self.data.org,
            type=Order.TYPE_CUSTOMER,
            applicant=applicant,
            dt=datetime.date.today(),
            customplace=self.data.customplace,
            status=Order.STATUS_ACCEPTED,
        )
        # будут назначены loru_number, number:
        order.save()

        if not self.data.customplace.favorite_performer:
            self.data.customplace.favorite_performer = self.data.org
            self.data.customplace.save()

        if self.data.service_name != Service.SERVICE_DELIVERY:
            item_service = ServiceItem.objects.create(
                order=order,
                orgservice=self.data.orgservice,
                cost=price_org,
            )
        if self.data.need_delivery:
            item_service = ServiceItem.objects.create(
                order=order,
                orgservice=self.data.orgservice_delivery,
                cost=price_delivery,
            )
            Route.objects.create(order=order, index=0, **closest_store)
            Route.objects.create(
                order=order,
                index=1,
                lat=self.data.latitude,
                lng=self.data.longitude,
            )
        comment = request.data.get('comment')
        if comment:
            OrderComment.objects.create(
                order=order,
                user=request.user,
                comment=comment,
            )
        #return Response(
            #data=dict(status='success', price=float(order.cost), currency=self.data.org.currency.code),
            #status=200,
        #)
        return Response(
            data=dict(
                id=order.pk,
                supplierId=self.data.org.pk,
                number=order.number_verbose(),
                type=self.data.service_name,
                placeId=self.data.customplace.pk,
                status=order.status,
            ),
            status=200,
        )

api_client_orders = ApiClientOrdersView.as_view()

class ApiServiceOrdersView(APIView):
    permission_classes = (PermitIfTradeOrCabinet,)

    def get(self, request):
        q = Q()
        if is_trade_user(request.user):
            q_type = Q(type=Order.TYPE_CUSTOMER) | \
                     Q(serviceitem__orgservice__service__name=Service.SERVICE_PHOTO)
            q = Q(loru=request.user.profile.org) & q_type
        elif is_cabinet_user(request.user):
            q = Q(customplace__user=request.user, type=Order.TYPE_CUSTOMER)
        qs = Order.objects.filter(q).order_by('-dt_created').distinct() if q else Order.objects.none()
        return Response(
            data=ServiceOrderSerializer(
                qs,
                many=True,
                context=dict(request=request),
                ).data,
            status=200
        )

api_orders = ApiServiceOrdersView.as_view()

class ApiOrderMixin(object):

    def get_order(self, pk):
        try:
            order = Order.objects.get(pk=pk)
            if not order.is_accessible(self.request.user):
                raise Http404
        except Order.DoesNotExist:
            raise Http404
        return order

class ApiOrderCommentsView(ApiOrderMixin, APIView):
    permission_classes = (PermitIfTradeOrCabinet,)

    def get(self, request, pk):
        order = self.get_order(pk=pk)
        data=[OrderCommentsSerializer(ordercomment, context=dict(
                request=request,
              )).data for ordercomment in OrderComment.objects.filter(order=order).order_by('dt_created')
        ]
        return Response(data=data, status=200)

    @transaction.atomic
    def post(self, request, pk):
        order = self.get_order(pk=pk)
        comment = request.data.get('comment')
        if comment is not None:
            ordercomment = OrderComment.objects.create(
                order=order,
                user=request.user,
                comment=comment,
            )

            if is_cabinet_user(request.user):
                email_to = order.loru.email
                if email_to:
                    order_url = "%s/%s" % (
                        get_host_url(request).rstrip('/'),
                        reverse('order_edit', args=[order.pk]).lstrip('/'),
                    )
                    profile = request.user.customerprofile
                    org = None
            else:
                # trade_user
                email_to = order.applicant and order.applicant.user and order.applicant.user.email
                if email_to:
                    order_url = order.loru.shop_site
                    if order_url:
                        order_url = "%s/%s" % (order_url.rstrip('/'), order.pk)
                    profile = request.user.profile
                    org = profile.org
            if email_to:
                email_text = render_to_string(
                                'order_notification.txt',
                                dict(
                                    comment=ordercomment,
                                    order_url=order_url,
                                    profile=profile,
                                    org=org,
                ))
                email_subject = _("Комментарий к заказу № %s") % order.number_verbose()
                email_from = settings.DEFAULT_FROM_EMAIL
                headers = dict()
                if request.user.email:
                    headers['Reply-To'] = request.user.email
                EmailMessage(email_subject, email_text, email_from, (email_to, ),
                             headers=headers).send(fail_silently=True)

            return Response(OrderCommentsSerializer(ordercomment, context=dict(request=request)).data, status=200)
        else:
            return Response(data=dict(status='error', message='No comment at input'), status=400)

api_orders_comments = ApiOrderCommentsView.as_view()

class ApiOrderResultView(ApiOrderMixin, APIView):
    permission_classes = (PermitIfTradeOrCabinet,)
    parser_classes = (MultiPartParser,)

    def get(self, request, pk):
        order = self.get_order(pk=pk)
        return Response(data=[OrderResultsSerializer(resultfile, context=dict(request=request)).data \
                        for resultfile in ResultFile.objects.filter(order=order).order_by('date_of_creation')],
                    status=200)

    @transaction.atomic
    def post(self, request, pk):
        try:
            order = self.get_order(pk=pk)
            if not is_trade_user(request.user):
                raise PermissionDenied
            type_ = request.data.get('type')
            if not type_:
                raise ServiceException(_("Не задан тип загружаемого файла"))
            for tp in ResultFile.RESULT_TYPES:
                if type_ == tp[0]:
                    break
            else:
                raise ServiceException(_("Тип %s файла результата выполнения заказа не предусмотрен") % type_)
            if 'file' not in request.data:
                raise ServiceException(_("Не получен загружаемый файл 'file'"))
            uploaded_file = request.data['file']
            if type_ == ResultFile.TYPE_IMAGE:
                if uploaded_file.size > ResultFile.MAX_IMAGE_SIZE * 1024 * 1024:
                    raise ServiceException(_("Размер изображения превышает %d Мб") % ResultFile.MAX_IMAGE_SIZE)
                if not get_image(uploaded_file):
                    raise ServiceException(_("Загруженный файл не является изображением"))
            elif type_ == ResultFile.TYPE_VIDEO:
                if not is_video(uploaded_file):
                    raise ServiceException(_("Загруженный файл не является видео"))
            is_title = str_to_bool_or_None(request.data.get('isTitle'))
            if is_title is None:
                is_title = True
            resultfile = ResultFile.objects.create(
                bfile=uploaded_file,
                order=order,
                type=type_,
                creator=request.user,
                is_title=is_title,
            )
            # отметим изменение в order.dt_modified:
            order.save()
        except ServiceException as excpt:
            transaction.set_rollback(True)
            return Response(data=dict(status='error', message=excpt.args[0]), status=400)
        return Response(data=OrderResultsSerializer(resultfile, context=dict(request=request)).data, status=200)

api_orders_results = ApiOrderResultView.as_view()

class ApiOrderResultDetailView(ApiOrderMixin, APIView):
    permission_classes = (PermitIfTradeOrCabinet,)

    def get_object(self, pk, result_pk, check_trade_user=True):
        if check_trade_user and not is_trade_user(self.request.user):
            raise PermissionDenied
        order = self.get_order(pk=pk)
        return get_object_or_404(ResultFile, order=order, pk=result_pk)

    def get(self, request, pk, result_pk):
        resultfile = self.get_object(pk, result_pk, check_trade_user=False)
        return Response(
            data=OrderResultsSerializer(resultfile, context=dict(request=request)).data,
            status=200,
        )

    def put(self, request, pk, result_pk):
        resultfile = self.get_object(pk, result_pk)
        is_title = request.data.get('isTitle')
        if is_title is not None and resultfile.type == ResultFile.TYPE_IMAGE:
            resultfile.is_title = is_title
            resultfile.save()
        return Response(
            data=OrderResultsSerializer(resultfile, context=dict(request=request)).data,
            status=200,
        )

    def delete(self, request, pk, result_pk):
        self.get_object(pk, result_pk).delete()
        return Response(data={}, status=200)

api_orders_results_detail = ApiOrderResultDetailView.as_view()

class ApiServiceOrderDetailView(ApiOrderMixin, APIView):
    permission_classes = (PermitIfTradeOrCabinet,)

    def get(self, request, pk):
        order = self.get_order(pk=pk)
        return Response(
            data=ServiceOrderDetailSerializer(order, context=dict(request=request)).data,
            status=200,
        )

api_orders_detail = ApiServiceOrderDetailView.as_view()

class ApiServiceOrderPutView(ApiOrderMixin, OptOrderMixin, APIView):
    permission_classes = (PermitIfTradeOrCabinet,)

    @transaction.atomic
    def put(self, request, pk):
        try:
            order = self.get_order(pk=pk)
            kwargs = dict()

            status = request.data.get('status')
            if status:
                for st in Order.STATUS_TYPES:
                    if status == st[0]:
                        break
                else:
                    raise ServiceException(_("Статус %s не предусмотрен") % status)
                if order.status != status:
                    kwargs.update(dict(status=status))

            if 'clientRating' in request.data:
                kwargs.update(dict(applicant_approved=request.data['clientRating']))

            archived = request.data.get('isArchived')
            if archived is not None:
                kwargs.update(dict(archived=archived))

            product_items = request.data.get('products')
            if product_items is not None:
                if order.status not in (Order.STATUS_POSTED, Order.STATUS_ACCEPTED,):
                    raise ServiceException(_("Набор товаров/услуг можно изменять только в размещенном заказе"))
                for orderitem in order.orderitem_set.all():
                    orderitem.delete()
                loru = order.loru
                for item in product_items:
                    try:
                        product = Product.objects.get(pk=item['id'])
                    except Product.DoesNotExist:
                        raise ServiceException(_("Не найден товар/услуга, id = %s") % item['id'])
                    if product.loru != loru:
                        raise ServiceException(_("Товар/услуга, id = %s, - не от исполнителя заказа") % item['id'])
                    quantity = item.get('qty', 1.00)
                    OrderItem.objects.create(
                        order=order,
                        product=product,
                        quantity=quantity,
                        cost=product.price,
                    )

            if kwargs:
                for f in kwargs:
                    setattr(order, f, kwargs[f])
                order.save()
                if kwargs.get('status'):
                    self.email_notifications(
                        order=order,
                        is_new_opt_order=False,
                        new_status=True,
                    )
            return Response(
                data=ServiceOrderDetailSerializer(order, context=dict(request=request)).data,
                status=200,
            )

        except ServiceException as excpt:
            transaction.set_rollback(True)
            return Response(data=dict(status='error', message=excpt.args[0]), status=400)

    def delete(self, request, pk):
        order = self.get_order(pk=pk)
        order.delete()
        return Response(data=dict(status='success'), status=200)

api_client_orders_put_status = ApiServiceOrderPutView.as_view()

class ApiOrderPaymentsMixin(object):

    PAY_SYSTEM_WEBPAY = 'webpay'
    PAY_SYSTEM_TYPES = (PAY_SYSTEM_WEBPAY, )

    def check_order_pay_system(self, order_pk, pay_system):
        if not pay_system:
            raise ServiceException(_("Не задан тип платежной системы"))
        if pay_system not in self.PAY_SYSTEM_TYPES:
            raise ServiceException(_("Платежная система %s не предусмотрена") % pay_system)
        try:
            order = Order.objects.get(pk=order_pk)
        except Order.DoesNotExist:
            raise Http404
        if not order.is_accessible(self.request.user):
            raise Http404
        system_not_supported = _("Исполнитель заказа, %s, не поддерживает платежи в системе %s")
        if pay_system == self.PAY_SYSTEM_WEBPAY:
            try:
                pay_data = OrgWebPay.objects.get(org=order.loru)
            except OrgWebPay.DoesNotExist:
                raise ServiceException(system_not_supported % (order.loru, pay_system,))
        return order, pay_data

class ApiOrderPaymentsView(ApiOrderPaymentsMixin, APIView):
    permission_classes = (PermitIfCabinet,)

    def get(self, request, pk, pay_system):
        try:
            order, pay_data = self.check_order_pay_system(order_pk=pk, pay_system=pay_system)
            if pay_system == self.PAY_SYSTEM_WEBPAY:
                orgwebpay = pay_data
                items = []
                for serviceitem in ServiceItem.objects.filter(order=order).order_by('pk'):
                    items.append(dict(
                        name=serviceitem.orgservice.service.title,
                        quantity="1.00",
                        price=str(serviceitem.cost)
                    ))
                for orderitem in OrderItem.objects.filter(order=order).order_by('name'):
                    items.append(dict(
                        name=orderitem.name,
                        quantity=str(orderitem.quantity),
                        price=str(orderitem.cost),
                    ))
                for key in ('quantity', 'price',):
                    for item in items:
                        if not item[key].endswith('.00'):
                            break
                    else:
                        for item in items:
                            item[key] = re.sub(r'\.00$','', item[key])
                data = dict(
                    wsb_storeid=orgwebpay.wsb_storeid,
                    wsb_store=orgwebpay.wsb_store,
                    wsb_order_num=order.number_webpay(),
                    wsb_currency_id=orgwebpay.wsb_currency_id,
                    wsb_version=orgwebpay.wsb_version,
                    wsb_language_id='russian',
                    wsb_seed=''.join(random.choice(string.ascii_uppercase + string.digits) for x in range(10)),
                    wsb_notify_url=get_host_url(self.request) + \
                                   reverse('api_orders_webpay_notify', args=(order.pk,)).strip('/'),
                    wsb_test="1" if orgwebpay.wsb_test else "0",
                    items=items,
                    wsb_total=str(order.total),
                )
                signature = hashlib.sha1() if data['wsb_version'] == "2" else hashlib.md5()
                signature.update(''.join((
                    data['wsb_seed'],
                    data['wsb_storeid'],
                    data['wsb_order_num'],
                    data['wsb_test'],
                    data['wsb_currency_id'],
                    data['wsb_total'],
                    orgwebpay.secret,
                )))
                data['wsb_signature'] = signature.hexdigest()
                OrderWebPay.objects.create(
                    order=order,
                    wsb_order_num=data['wsb_order_num'],
                )
        except ServiceException as excpt:
            return Response(data=dict(status='error', message=excpt.args[0]), status=400)
        return Response(data=data, status=200)

api_orders_payments = ApiOrderPaymentsView.as_view()

class ApiWebPayNotifyView(APIView):
    parser_classes = (FormParser,)
    renderer_classes = (StaticHTMLRenderer, )

    @transaction.atomic
    def post(self, request, pk):
        try:
            order = Order.objects.get(pk=pk)
        except Order.DoesNotExist:
            raise Http404
        try:
            orderwebpay = OrderWebPay.objects.filter(order=order).order_by('-pk')[0]
        except IndexError:
            raise Http404

        # номер заказа приходит в POST-параметре 'site_order_id',
        # а не в GET параметре wsb_order_num, как описано в документации webpay
        #
        post_keys = (
            'transaction_id',
            'batch_timestamp',
            'currency_id',
            'amount',
            'payment_method',
            'payment_type',
            'order_id',
            'rrn',
            'wsb_signature',
        )

        input_data = dict()
        for post_key in post_keys:
            model_key = 'order_ident' if post_key == 'order_id' else post_key
            input_data[model_key] = request.data.get(post_key)

        if input_data['payment_type'] in OrderWebPay.SUCCESS_PAY_TYPES:
            order.status = Order.STATUS_PAID
            order.save()

        for key in input_data:
            setattr(orderwebpay, key, input_data[key])
            orderwebpay.save()
        return Response('', status=200)

api_orders_webpay_notify = ApiWebPayNotifyView.as_view()

class ApiClientOrderPaymentsView(ApiOrderPaymentsMixin, APIView):
    permission_classes = (PermitIfCabinet,)

    @transaction.atomic
    def post(self, request, pk):
        """
        Отметка заказа как оплаченного

        На входе:
            type: тип платежносй системы
            paymentToken - номер транзакции, который платежная система возвращает во front-end
            в случае успешного платежа, и который front-end будет передавать в апи для валидации
            и последующей смены статуса заказа на Оплаченый, в случае успешной валидации.

            Для webpay - это атрибут wsb_tid. Не смотря на то, что сервер получает wsb_notify_url
            front-end будет передавать в апи и этот запрос, на случай если wsb_notify_url
            не достучался до сервера или по др. причине, даже если ранее апи уже выставило
            статус для заказа paid, при получении этого запроса - должен будет вернуть 201,
            либо если еще не выставило, то выставить и вернуть 201
        """
        pay_system = request.data.get('type')
        try:
            order, pay_data = self.check_order_pay_system(order_pk=pk, pay_system=pay_system)
            if pay_system == self.PAY_SYSTEM_WEBPAY:
                transaction_id = request.data.get('paymentToken')
                if not transaction_id:
                    raise ServiceException(_("Не задан параметр paymentToken (идентификатор транзакции webpay)"))
            try:
                orderwebpay = OrderWebPay.objects.filter(order=order).order_by('-pk')[0]
            except IndexError:
                raise ServiceException(_("Не была выставлена оплата за заказ"))
            if not orderwebpay.transaction_id:
                orderwebpay.transaction_id=transaction_id
                orderwebpay.save()
            if order.status != Order.STATUS_PAID:
                order.status = Order.STATUS_PAID
                order.save()
        except ServiceException as excpt:
            transaction.set_rollback(True)
            return Response(data=dict(status='error', message=excpt.args[0]), status=400)
        return Response(data={}, status=201)

api_client_orders_payments = ApiClientOrderPaymentsView.as_view()

class OrderItemCheckMixin(object):

    def checked_order_item(self, request, item):
        product_id = item.get('id')
        if not product_id:
            raise ServiceException(_('Нет productId'))

        try:
            product = Product.objects.get(loru=request.user.profile.org, pk=product_id)
        except Product.DoesNotExist:
            raise ServiceException(
                _('Не найден productId=%s вообще или у этого поставщика') % product_id)

        msg = _('Неверное количество у продукта/услуги: %s') % product.name
        try:
            quantity = decimal.Decimal(item.get('amount') or 1)
        except decimal.InvalidOperation:
            raise ServiceException(msg)
        if quantity < 0:
            raise ServiceException(msg)

        msg = _('Неверная скидка у продукта/услуги: %s') % product.name
        try:
            discount = decimal.Decimal(item.get('discount') or 0)
        except decimal.InvalidOperation:
            raise ServiceException(msg)
        if discount < 0 or discount > 100:
            raise ServiceException(msg)

        return product, quantity, discount

class ApiLoruOrdersView(
    CheckLifeDatesMixin,
    UnclearDateFieldMixin,
    TradeCemeteriesMixin,
    OrderItemCheckMixin,
    APIView
):
    permission_classes = (PermitIfTrade,)

    @transaction.atomic
    def post(self, request):
        msg_invalid_dt_due = _("Неверная дата похорон")
        msg_invalid_dt = _("Неверная дата создания заказа")
        try:
            customer = request.data.get('customer')
            login_phone = None
            applicant = None
            if customer:
                if customer.get('createCabinet'):
                    login_phone_str = customer.get('phoneNumber', '').strip()
                    if not login_phone_str:
                        raise ServiceException(_("Создать кабинет: не указан телефон"))
                    try:
                        validate_phone_as_number(login_phone_str)
                    except ValidationError as excpt:
                        raise ServiceException(
                            "Создать кабинет. %s" % \
                                (excpt.args and excpt.args[0] or _('Не удалось по неизвестной причине'))
                        )
                    login_phone = decimal.Decimal(login_phone_str)
                if customer.get('address'):
                    address = Location.objects.create(addr_str=customer['address'])
                else:
                    address = None
                applicant = AlivePerson.objects.create(
                    last_name=customer.get('lastName', ''),
                    first_name=customer.get('firstName', ''),
                    middle_name=customer.get('middleName', ''),
                    phones=customer.get('phoneNumber', ''),
                    address=address,
                    login_phone=login_phone,
                )
            dt_due = request.data.get('dueDate') or None
            if dt_due:
                try:
                    dt_due = datetime.datetime.strptime(dt_due, "%d.%m.%Y").date()
                except ValueError:
                    raise ServiceException(msg_invalid_dt_due)
                try:
                    dt_due.strftime("%d.%m.%Y")
                except ValueError:
                    raise ServiceException(msg_invalid_dt_due)
            dt = request.data.get('createdDate') or None
            mapping = dict(
                burialPlanTime='burial_plan_time',
                initialTime='initial_time',
                serviceTime='service_time',
                repastTime='repast_time',
            )
            other_keys = dict()
            for k in mapping:
                if request.data.get(k):
                    try:
                        other_keys[mapping[k]] = datetime.datetime.strptime(request.data[k], '%H:%M').time()
                    except ValueError:
                        raise ServiceException(_("Неверное %s") %
                            Order._meta.get_field(mapping[k]).verbose_name.lower()
                        )
            mapping = dict(
                initialPlace='initial_place',
                servicePlace='service_place',
                repastPlace='repast_place',
            )
            for k in mapping:
                if request.data.get(k):
                    other_keys[mapping[k]] = request.data[k]
            if dt:
                try:
                    dt = datetime.datetime.strptime(dt, "%d.%m.%Y").date()
                except ValueError:
                    raise ServiceException(msg_invalid_dt)
                try:
                    dt.strftime("%d.%m.%Y")
                except ValueError:
                    raise ServiceException(msg_invalid_dt)
            order = Order.objects.create(
                loru=request.user.profile.org,
                applicant=applicant,
                dt=dt or datetime.date.today(),
                dt_due=dt_due,
                type=Order.TYPE_FUNERAL,
                **other_keys
            )
            deadman = request.data.get('deadman')
            if deadman:
                message = self.check_life_dates(person=deadman, format='d.m.y')
                if message:
                    raise ServiceException(message)
                deadman = OrderDeadPerson.objects.create(
                    order=order,
                    last_name=deadman.get('lastName',''),
                    first_name=deadman.get('firstName',''),
                    middle_name=deadman.get('middleName',''),
                    birth_date=self.set_unclear_date(deadman.get('dob'), format='d.m.y'),
                    death_date=self.set_unclear_date(deadman.get('dod'), format='d.m.y'),
                )
            place = request.data.get('place')
            if place:
                cemetery = None
                if place.get('cemeteryId') or place.get('areaId'):
                    cemeteries = self.available_cemeteries(request.user)
                if place.get('cemeteryId'):
                    cemetery_msg = _("Нет такого кладбища среди доступных")
                    try:
                        cemetery = Cemetery.objects.get(pk=place['cemeteryId'])
                    except Cemetery.DoesNotExist:
                        raise ServiceException(cemetery_msg)
                    if cemetery not in cemeteries:
                        raise ServiceException(cemetery_msg)
                area = None
                if place.get('areaId'):
                    area_msg = _("Нет такого участка на доступных кладбищах")
                    try:
                        area = Area.objects.get(pk=place['areaId'])
                    except Area.DoesNotExist:
                        raise ServiceException(area_msg)
                    cemetery = area.cemetery
                    if cemetery not in cemeteries:
                        raise ServiceException(area_msg)
                cemetery_text = place.get('cemeteryText', '')
                row = place.get('row', '')
                place_number = place.get('placeNumber', '')
                size = place.get('size', '')
                place = OrderPlace.objects.create(
                    order=order,
                    cemetery=cemetery,
                    area=area,
                    cemetery_text=cemetery_text,
                    row=row,
                    place=place_number,
                    size=size,
                )
            products = request.data.get('products', [])
            for item in products:
                product, quantity, discount = self.checked_order_item(request, item)
                orderitem = OrderItem.objects.create(
                    order=order,
                    product=product,
                    quantity=quantity,
                    discount=discount,
                )
            debug_text = None
            if login_phone:
                text = None
                try:
                    customerprofile = CustomerProfile.objects.get(login_phone=login_phone)
                    user = customerprofile.user
                    debug_text = _("Заказ похорон с созданием кабинета. Пользователь %s (телефон %s) уже существует") % \
                                (user.username, login_phone)
                    write_log(request, order, debug_text)
                except CustomerProfile.DoesNotExist:
                    user, password = CustomerProfile.create_cabinet(applicant, request)
                    text = _('%s login: %s parol: %s') % (
                        get_front_end_url(request).rstrip('/'),
                        login_phone,
                        password,
                    )
                    email_error_text = _("Пользователь %s не смог получить пароль по СМС после создания заказа похорон") % \
                                        login_phone
                    debug_text = _("Создан кабинет. Пользователь %s, пароль %s") % \
                                (user.username, password)
                    if not settings.DEBUG:
                        sent, message = send_sms(
                            phone_number=login_phone,
                            text=text,
                            email_error_text=email_error_text,
                            user=request.user,
                        )
        except ServiceException as excpt:
            transaction.set_rollback(True)
            return Response(data=dict(status='error', message=excpt.args[0]), status=400)
        write_log(request, order, _('Заказ создан'))
        data = LoruOrderSerializer(order, context=dict(request=request)).data
        if settings.DEBUG and debug_text:
            data.update(dict(debugMessage=debug_text))
        return Response(data=data, status=200)

api_loru_orders = ApiLoruOrdersView.as_view()

class ApiLoruCategoriesView(APIView):
    permission_classes = (PermitIfTrade,)

    def get(self, request):
        return Response(
            data=[ProductCategory2Serializer(
                    category,
                    context=dict(request=request),
                ).data for category in ProductCategory.objects. \
                             filter(product__loru=request.user.profile.org, product__is_archived=False). \
                             order_by('sorting', 'name').distinct()
            ],
            status=200,
        )

api_loru_categories = ApiLoruCategoriesView.as_view()

class ApiLoruOrdersDetailView(
    SafeDeleteMixin,
    CheckLifeDatesMixin,
    UnclearDateFieldMixin,
    TradeCemeteriesMixin,
    OrderItemCheckMixin,
    APIView
):
    permission_classes = (PermitIfTrade,)

    def get(self, request, pk):
        order = get_object_or_404(Order, loru=request.user.profile.org, pk=pk)
        return Response(
            data=LoruOrderSerializer(order,
                context=dict(request=request),
                ).data,
            status=200,
        )

    @transaction.atomic
    def put(self, request, pk):
        order = get_object_or_404(Order, loru=request.user.profile.org, pk=pk)
        msg_invalid_dt_due = _("Неверная дата похорон")
        msg_invalid_dt = _("Неверная дата создания заказа")
        try:
            order_save = False
            if 'dueDate' in request.data:
                dt_due = request.data['dueDate'] or None
                if dt_due:
                    try:
                        dt_due = datetime.datetime.strptime(dt_due, "%d.%m.%Y").date()
                    except ValueError:
                        raise ServiceException(msg_invalid_dt_due)
                    try:
                        dt_due.strftime("%d.%m.%Y")
                    except ValueError:
                        raise ServiceException(msg_invalid_dt_due)
                if order.dt_due != dt_due:
                    order.dt_due = dt_due
                    order_save = True
            dt = request.data.get('createdDate') or None
            if dt:
                try:
                    dt = datetime.datetime.strptime(dt, "%d.%m.%Y").date()
                except ValueError:
                    raise ServiceException(msg_invalid_dt)
                try:
                    dt.strftime("%d.%m.%Y")
                except ValueError:
                    raise ServiceException(msg_invalid_dt)
                if order.dt != dt:
                    order.dt = dt
                    order_save = True
            mapping = dict(
                burialPlanTime='burial_plan_time',
                initialTime='initial_time',
                serviceTime='service_time',
                repastTime='repast_time',
            )
            other_keys = dict()
            for k in mapping:
                if k in request.data:
                    f = request.data[k] and request.data[k].strip() or None
                    if f is not None:
                        try:
                            f = datetime.datetime.strptime(f, '%H:%M').time()
                        except ValueError:
                            raise ServiceException(_("Неверное %s") %
                                Order._meta.get_field(mapping[k]).verbose_name.lower()
                            )
                    if f != getattr(order, mapping[k]):
                        order_save = True
                        setattr(order, mapping[k], f)
            mapping = dict(
                initialPlace='initial_place',
                servicePlace='service_place',
                repastPlace='repast_place',
            )
            for k in mapping:
                if k in request.data:
                    f = request.data[k] and request.data[k].strip() or ''
                    if f != getattr(order, mapping[k]):
                        order_save = True
                        setattr(order, mapping[k], f)
            if 'customer' in request.data:
                customer = request.data['customer']
                if customer is None:
                    self.safe_delete('applicant', order)
                else:
                    if order.applicant:
                        applicant = order.applicant
                        mapping = (
                            ('lastName', 'last_name',),
                            ('firstName', 'first_name',),
                            ('middleName', 'middle_name',),
                            ('phoneNumber', 'phones',),
                        )
                        applicant_save = False
                        for req, field in mapping:
                            if req in customer and getattr(applicant, field) != customer[req]:
                                applicant_save = True
                                setattr(applicant, field, customer[req])
                        if customer.get('address'):
                            if applicant.address:
                                if applicant.address.addr_str != customer['address']:
                                    applicant.address.addr_str = customer['address']
                                    applicant.address.save()
                            else:
                                applicant.address = Location.objects.create(addr_str=customer['address'])
                                applicant_save = True
                        elif 'address' in customer and applicant.address:
                            self.safe_delete('address', applicant)
                        if applicant_save:
                            applicant.save()
                    else:
                        if customer.get('address'):
                            address = Location.objects.create(addr_str=customer['address'])
                        else:
                            address = None
                        applicant = AlivePerson.objects.create(
                            last_name=customer.get('lastName', ''),
                            first_name=customer.get('firstName', ''),
                            middle_name=customer.get('middleName', ''),
                            phones=customer.get('phoneNumber', ''),
                            address=address,
                        )
                        order.applicant = applicant
                        order_save = True
            if 'products' in request.data:
                products = request.data['products'] or []
                for item in OrderItem.objects.filter(order=order):
                    item.delete()
                for item in products:
                    product, quantity, discount = self.checked_order_item(request, item)
                    orderitem = OrderItem.objects.create(
                        order=order,
                        product=product,
                        quantity=quantity,
                        discount=discount,
                    )
            if 'deadman' in request.data:
                deadman = request.data['deadman']
                try:
                    instance = order.orderdeadperson
                except (OrderDeadPerson.DoesNotExist, AttributeError,):
                    instance = None
                if deadman is not None:
                    message = self.check_life_dates(
                        person=deadman,
                        instance=instance,
                        format='d.m.y',
                    )
                    if message:
                        raise ServiceException(message)
                if instance:
                    if deadman is None:
                        instance.delete()
                    else:
                        mapping = (
                            ('lastName', 'last_name',),
                            ('firstName', 'first_name',),
                            ('middleName', 'middle_name',),
                        )
                        deadman_save = False
                        for req, field in mapping:
                            if req in deadman and getattr(instance, field) != deadman[req]:
                                deadman_save = True
                                setattr(instance, field, deadman[req])
                        if 'dob' in deadman:
                            instance.birth_date = self.set_unclear_date(deadman['dob'], format='d.m.y')
                            deadman_save = True
                        if 'dod' in deadman:
                            instance.death_date = self.set_unclear_date(deadman['dod'], format='d.m.y')
                            deadman_save = True
                        if deadman_save:
                            instance.save()
                elif deadman is not None:
                    deadman = OrderDeadPerson.objects.create(
                        order=order,
                        last_name=deadman.get('lastName',''),
                        first_name=deadman.get('firstName',''),
                        middle_name=deadman.get('middleName',''),
                        birth_date=self.set_unclear_date(deadman.get('dob'), format='d.m.y'),
                        death_date=self.set_unclear_date(deadman.get('dod'), format='d.m.y'),
                    )
            if 'place' in request.data:
                place = request.data['place']
                cemetery = area = None
                if place is not None:
                    if place.get('cemeteryId') or place.get('areaId'):
                        cemeteries = self.available_cemeteries(request.user)
                    if place.get('cemeteryId'):
                        cemetery_msg = _("Нет такого кладбища среди доступных")
                        try:
                            cemetery = Cemetery.objects.get(pk=place['cemeteryId'])
                        except Cemetery.DoesNotExist:
                            raise ServiceException(cemetery_msg)
                        if cemetery not in cemeteries:
                            raise ServiceException(cemetery_msg)
                    if place.get('areaId'):
                        area_msg = _("Нет такого участка на доступных кладбищах")
                        try:
                            area = Area.objects.get(pk=place['areaId'])
                        except Area.DoesNotExist:
                            raise ServiceException(area_msg)
                        cemetery = area.cemetery
                        if cemetery not in cemeteries:
                            raise ServiceException(area_msg)
                try:
                    instance = order.orderplace
                except (OrderPlace.DoesNotExist, AttributeError,):
                    instance = None
                if instance:
                    if place is None:
                        instance.delete()
                    else:
                        place_save = False
                        mapping = (
                            ('cemeteryText', 'cemetery_text',),
                            ('row', 'row',),
                            ('placeNumber', 'place',),
                            ('size', 'size',),
                        )
                        for req, field in mapping:
                            if req in place and getattr(instance, field) != place[req]:
                                place_save = True
                                setattr(instance, field, place[req])
                        if ('cemeteryId' in place or cemetery) and cemetery != instance.cemetery:
                            place_save = True
                            instance.cemetery = cemetery
                        if 'areaId' in place and area != instance.area:
                            place_save = True
                            instance.area = area
                        if place_save:
                            instance.save()
                elif place is not None:
                    place = OrderPlace.objects.create(
                        order=order,
                        cemetery=cemetery,
                        area=area,
                        cemetery_text=place.get('cemeteryText', ''),
                        row=place.get('row', ''),
                        place=place.get('placeNumber', ''),
                        size=place.get('size', ''),
                    )
            if order_save:
                order.save()
        except ServiceException as excpt:
            transaction.set_rollback(True)
            return Response(data=dict(status='error', message=excpt.args[0]), status=400)
        write_log(request, order, _('Заказ изменен'))
        return Response(
            data=LoruOrderSerializer(order,
                context=dict(request=request),
                ).data,
            status=200,
        )

api_loru_orders_detail = ApiLoruOrdersDetailView.as_view()

class ProductXlsxreportView(LORURequiredMixin, FormView):

    template_name = 'product_xlsxreport.html'
    form_class = ProductXlsxreportForm

    def get_form(self, *args, **kwargs):
        form = super(ProductXlsxreportView, self).get_form(*args, **kwargs)
        today = datetime.date.today()
        form.initial['date_from'] = today
        form.initial['date_to'] = today
        return form

    def thousand_str(self, num):
        return  num # f"{num:,}".replace(',', ' ')

    def form_valid(self, form, *args, **kwargs):

        org = self.request.user.profile.org
        org_pk = org.pk
        media_path = os.path.join('tmp', 'products', 'reports',
            '%s' % org_pk,
        )
        export_path = os.path.join(settings.MEDIA_ROOT, media_path)
        try:
            os.makedirs(export_path)
        except OSError:
            pass
        temp_dir = tempfile.mkdtemp(dir=export_path)
        temp_dir_name = os.path.basename(temp_dir)
        d = datetime.datetime.now()
        date_from = form.cleaned_data['date_from']
        date_from_str = datetime.datetime.strftime(date_from, '%Y%m%d')
        date_to = form.cleaned_data['date_to']
        date_to_str = datetime.datetime.strftime(date_to, '%Y%m%d')
        date_to_1 = date_to + datetime.timedelta(days=1)
        fname_basename = 'report-from-%s-to-%s.xlsx' % (date_from_str, date_to_str,)
        fname = os.path.join(temp_dir, fname_basename)
        
        first = True
        row = 0
        npp = 0
        total = 0
        qs = OrderItem.objects.filter(
                    order__loru=org,
                    order__dt__gte=date_from,
                    order__dt__lt=date_to_1
                ).order_by('product__name'). \
                values(
                    'product__name',
                    'product__measure',
                    'product__sku',
                ). \
                annotate(
                    sum=Sum(F('cost') * F('quantity') * (100 - F('discount')) / 100),
                    count=Sum(F('quantity'))
                )
        qs_count = qs.count()
        for p in qs:
            if first:
                row = 2
                thick = Side(border_style="thick")
                thin = Side(border_style="thin")
                align_right = Alignment(horizontal="right", vertical="center", indent=1)
                align_left = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
                book = Workbook()
                sheet = book.active
                sheet.title = 'Отчет'
                if date_from == date_to:
                    title_end = 'за %s' % datetime.datetime.strftime(date_from, '%m.%d.%Y')
                else:
                    title_end = 'с %s по %s' % (
                        datetime.datetime.strftime(date_from, '%m.%d.%Y'),
                        datetime.datetime.strftime(date_to, '%m.%d.%Y'),
                    )
                title = sheet.cell(row, 1)
                title.value = 'Отчет о розничных продажах %s' % title_end
                for c in range(1, 7):
                    sheet.cell(row, c).border = Border(bottom=thick)
                title.font  = Font(name='Arial', b=True, size=16)
                row = 4
                title = sheet.cell(row, 1)
                title.value = 'Организация: %s' % org.name
                title.font  = Font(b=True, size=14)
                row = 6
                rd = sheet.row_dimensions[row]
                rd.height = 25
                sheet.cell(row, 1).value = '№'
                sheet.cell(row, 2).value = 'Артикул'
                sheet.cell(row, 3).value = 'Товар'
                sheet.merge_cells('D%s:E%s' % (row, row,))
                top_left_cell = sheet['D%s' % row]
                top_left_cell.value = 'Количество'
                sheet.cell(row, 6).value = 'Цена'
                for col in range(1,7):
                    sheet.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
                    sheet.cell(row, col).font  = Font(name='Arial', b=True, size=12)
                    if col == 1:
                        sheet.cell(row, col).border = Border(top=thick, left=thick, right=thin, bottom=thick)
                    elif col == 6:
                        sheet.cell(row, col).border = Border(top=thick, left=thin, right=thick, bottom=thick)
                    else:
                        sheet.cell(row, col).border = Border(top=thick, left=thin, right=thin, bottom=thick)
                sheet.column_dimensions['A'].width =  7
                sheet.column_dimensions['B'].width = 10
                sheet.column_dimensions['C'].width = 35
                sheet.column_dimensions['D'].width = 7
                sheet.column_dimensions['E'].width = 10
                sheet.column_dimensions['F'].width = 14
                first = False
            row += 1
            npp += 1
            cell_npp = sheet.cell(row, 1)
            cell_npp.value = str(npp)
            # Если здесь вставить align_right, то некоторые знаки здесь криво ???
            cell_npp.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True, indent=1)

            cell_sku = sheet.cell(row, 2)
            cell_sku.value = str(p['product__sku'])
            cell_sku.alignment = align_left

            cell_name = sheet.cell(row, 3)
            cell_name.value = p['product__name']
            cell_name.alignment = align_left

            v_count = p['count']
            if int(v_count) == v_count:
                v_count = int(v_count)
            cell_count = sheet.cell(row, 4)
            cell_count.value = str(int(v_count))
            cell_count.alignment = align_right

            cell_measure = sheet.cell(row, 5)
            cell_measure.value = p['product__measure']
            cell_measure.alignment = align_left

            v_sum = round(p['sum'], 2)
            total += v_sum
            cell_sum = sheet.cell(row, 6)
            cell_sum.value = self.thousand_str(v_sum)
            cell_sum.alignment = align_right

            if npp == qs_count:
                border_bottom = thick
            else:
                border_bottom = thin
            for col in range(1, 7):
                if col == 1:
                    sheet.cell(row, col).border = Border(left=thick, right=thin, bottom=border_bottom)
                elif col == 6:
                    sheet.cell(row, col).border = Border(left=thin, right=thick, bottom=border_bottom)
                else:
                    sheet.cell(row, col).border = Border(left=thin, right=thin, bottom=border_bottom)
                sheet.cell(row, col).font = Font(name='Arial', size=10)

            row_height = sheet.row_dimensions[row].height = 15
            col_width = sheet.column_dimensions['C'].width
            mul = math.ceil(1.1 * len(str(cell_name.value)) / col_width)
            if mul > 1:
                sheet.row_dimensions[row].height = int(row_height * mul)
        
        if row:
            row += 1
            itogo = sheet.cell(row, 5)
            itogo.value = 'Итого:'
            itogo.font = Font(name='Arial', b=True, size=12)
            itogo.alignment = align_right

            itogo_sum = sheet.cell(row, 6)
            itogo_sum.value = self.thousand_str(total)
            itogo_sum.font = Font(name='Arial', b=True, size=10)
            itogo_sum.alignment = align_right

            row += 1
            v_n = sheet.cell(row, 1)
            v_n.value = 'Всего наименований %s, на сумму %s руб.' % (
                npp, itogo_sum.value
            )
            row += 1
            sheet.row_dimensions[row].height = 20
            propis = sheet.cell(row, 1)
            kopecks = '%02d' % int((total - int(total)) * 100)
            roubles = pytils.numeral.in_words(int(total))
            roubles = roubles[0].upper() + roubles[1:]
            propis.value = '%s руб. %s коп.' % (roubles, kopecks)
            propis.font = Font(name='Arial', b=True, size=10)

            row += 1
            for c in range(1, 7):
                sheet.cell(row, c).border = Border(bottom=thick)

            book.save(fname)
            return redirect(os.path.join(settings.MEDIA_URL, media_path, temp_dir_name, fname_basename))
        else:
            try:
                os.rmdir(temp_dir)
            except OSError:
                pass
            messages.info(self.request, _('Не найдены данные для отчета за указанный интервал дат'))
            return self.get(self.request, *self.args, **self.kwargs)

product_xlsxreport = ProductXlsxreportView.as_view()
